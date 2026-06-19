"""
[한솔테크닉스 HEVH] LOSSTIME + SCRAP 분석 대시보드 v3.9
- 헤더 클릭 → 초기화면(필터 리셋)
- 수치 소수점 1자리
- 공정별 파이차트 클릭 → 손실유형 바차트
- GitHub 자동 저장/로드
실행: python -m streamlit run dashboard.py
"""

import re
import io
import os
import base64
import requests
import pandas as pd
import streamlit as st
import plotly.express as px
import plotly.graph_objects as go
from openpyxl import load_workbook
from openpyxl.chartsheet import Chartsheet

st.set_page_config(
    page_title="HEVH 분석 대시보드",
    page_icon="🏭",
    layout="wide",
    initial_sidebar_state="expanded"
)

st.markdown("""
<style>
.main-header {
    background: linear-gradient(90deg, #1e3a5f, #2563eb);
    padding: 20px; border-radius: 10px; margin-bottom: 20px;
    color: white; text-align: center;
    cursor: pointer;
}
.main-header:hover { opacity: 0.9; }
.metric-box {
    background: #f8fafc; border: 1px solid #e2e8f0;
    border-radius: 8px; padding: 16px; text-align: center;
    margin-bottom: 8px;
}
.metric-box:hover { background:#e8f4fd; border-color:#2563eb; }
.metric-val { font-size: 28px; font-weight: bold; color: #1e3a5f; }
.metric-lbl { font-size: 13px; color: #64748b; margin-top: 4px; }
.pm-p1 { background:#fef2f2; border-left:4px solid #dc2626;
          padding:10px; border-radius:4px; margin:4px 0; }
.pm-p2 { background:#fff7ed; border-left:4px solid #f97316;
          padding:10px; border-radius:4px; margin:4px 0; }
.pm-p3 { background:#fefce8; border-left:4px solid #eab308;
          padding:10px; border-radius:4px; margin:4px 0; }
.detail-box {
    background:#f0f9ff; border:1px solid #bae6fd;
    border-radius:8px; padding:12px; margin-top:8px;
}
</style>
""", unsafe_allow_html=True)

DAY_SLOTS   = ["A","B","C","D","E","F"]
NIGHT_SLOTS = ["F","G","H","I","J","K"]
DB_PATH     = "losstime_db.csv"
SCRAP_DB    = "scrap_db.csv"

TYPE_COLOR = {
    "모델교체":           "#dc2626",
    "Magazine부족":       "#f97316",
    "신규OP교육":         "#f59e0b",
    "Printer불량":        "#84cc16",
    "SMT대기":            "#06b6d4",
    "기타":               "#8b5cf6",
    "자재부족":           "#ec4899",
    "RH3삽입불량":        "#14b8a6",
    "Mouter불량":         "#3b82f6",
    "Axial불량":          "#a855f7",
    "Wave Solder불량":    "#0891b2",
    "ICT/ATE불량":        "#65a30d",
    "AOI/S-AOI불량":      "#d97706",
    "계획완료":           "#6b7280",
    "설비고장(기타)":     "#ef4444",
    "Coating/Reflow불량": "#10b981",
    "XGZ불량":            "#7c3aed",
    "Jumper불량":         "#db2777",
    "SPI불량":            "#0369a1",
    "Inloader불량":       "#9ca3af",
}

LOSS_TYPE_RULES = [
    ("NEW_OP",       "신규OP교육",         ["new op","đào tạo","op mới"]),
    ("MAGAZINE",     "Magazine부족",       ["magazine","magazin","mag"]),
    ("WAITING_SMT",  "SMT대기",            ["waiting semi","đợi semi","chờ semi"]),
    ("MATERIAL",     "자재부족",           ["thiếu liệu","chờ liệu","đợi liệu",
                                            "thiếu hàng","chờ hàng","thiếu vật tư"]),
    ("CHANGE_MODEL", "모델교체",           ["change model","đổi model","tách lot",
                                            "cover work","đổi ca","chuyển model"]),
    ("PRINTER",      "Printer불량",        ["printer","priter","lỗi keo","tràn keo"]),
    ("MOUTER",       "Mouter불량",         ["moutor","mounter","stopper","băng tải"]),
    ("INSERT_RH3",   "RH3삽입불량",        ["rh3","rh 3","rhu"]),
    ("INSERT_XGZ",   "XGZ불량",            ["xzg","xgz"]),
    ("INSERT_AXIAL", "Axial불량",          ["axial","radial","cắm rớt","rad lỗi"]),
    ("INSERT_JUMP",  "Jumper불량",         ["jumper","jump "]),
    ("COATING",      "Coating/Reflow불량", ["coating","reflow","flux","nhiệt"]),
    ("AOI_SAOI",     "AOI/S-AOI불량",      ["saoi","s-aoi","aoi"]),
    ("WAVE",         "Wave Solder불량",    ["wave solder","wave"]),
    ("SPI",          "SPI불량",            ["spi"]),
    ("INLOADER",     "Inloader불량",       ["inloader"]),
    ("ICT_ATE",      "ICT/ATE불량",        ["ict","ate","ft lỗi"]),
    ("EQUIP_FAIL",   "설비고장(기타)",     ["hư","lỗi máy","stop line","spare"]),
    ("DONE_PLAN",    "계획완료",           ["done plan","hết plan","kết thúc"]),
    ("SAMPLE",       "샘플/테스트",        ["sample"]),
    ("ETC",          "기타",               []),
]

SCRAP_CAUSE_RULES = [
    ("BROKEN_DROP",  "낙하/파손",     ["rớt","rơi","bể","vỡ","broken","drop","rộng","sập"]),
    ("BROKEN_BI",    "Burn-in파손",   ["burn in","burin","tủ burin","ốc","tuột"]),
    ("SOLDER",       "납불량",        ["tràn keo","solder over","hàn","keo"]),
    ("CRACK",        "크랙/균열",     ["nứt","crack","bể mạch"]),
    ("PAD_DAMAGE",   "Pad손상",       ["tróc pad","pad đồng","tróc"]),
    ("EYELET",       "Eyelet불량",    ["eyelet","thiếu eyelet"]),
    ("SENSOR",       "센서불량",      ["sensor","cảm biến","chinh lai sensor"]),
    ("MAGAZINE",     "Magazine불량",  ["magazine","megazine","mag"]),
    ("NEW_OP",       "신규OP실수",    ["op mới","thời vụ","mới","new op"]),
    ("AUTO",         "자동판정(ATE)", ["atuo_scrap","auto_scrap"]),
    ("ETC",          "기타",          []),
]

OPER_MAP   = {"P-AUTO":"AI","P-SMD":"SMT","P-PBA":"MI"}
PROC_COLOR = {"AI":"#ef4444","SMT":"#3b82f6","MI":"#10b981","기타":"#9ca3af"}

# ─────────────────────────────────────────────
# GitHub DB 관리
# ─────────────────────────────────────────────
def get_github_config():
    try:
        return st.secrets["github"]["token"], st.secrets["github"]["repo"]
    except: return None, None

def github_load_csv(filename):
    token,repo = get_github_config()
    if not token:
        try: return pd.read_csv(filename, encoding="utf-8-sig")
        except FileNotFoundError: return pd.DataFrame()
    url = f"https://api.github.com/repos/{repo}/contents/{filename}"
    headers = {"Authorization":f"token {token}",
               "Accept":"application/vnd.github.v3+json"}
    try:
        r = requests.get(url, headers=headers, timeout=10)
        if r.status_code==200:
            return pd.read_csv(io.StringIO(
                base64.b64decode(r.json()["content"]).decode("utf-8")))
        return pd.DataFrame()
    except: return pd.DataFrame()

def github_save_csv(df, filename, msg=None):
    token,repo = get_github_config()
    if not token:
        df.to_csv(filename, index=False, encoding="utf-8-sig"); return True
    url = f"https://api.github.com/repos/{repo}/contents/{filename}"
    headers = {"Authorization":f"token {token}",
               "Accept":"application/vnd.github.v3+json"}
    sha = None
    try:
        r = requests.get(url, headers=headers, timeout=10)
        if r.status_code==200: sha = r.json().get("sha")
    except: pass
    csv_str = df.to_csv(index=False, encoding="utf-8-sig")
    content  = base64.b64encode(csv_str.encode("utf-8-sig")).decode()
    payload  = {"message":msg or f"DB업데이트:{filename}","content":content}
    if sha: payload["sha"] = sha
    try:
        r = requests.put(url, headers=headers, json=payload, timeout=15)
        return r.status_code in [200,201]
    except: return False

# ─────────────────────────────────────────────
# 로그인
# ─────────────────────────────────────────────
def check_login():
    try: return st.secrets["users"]
    except: return {"hansol":"hevh2024","vietnam":"hevh2024","korea":"hevh2024"}

def login_page():
    st.markdown("""
    <div style="text-align:center;margin-top:60px;">
        <h2>🏭 한솔테크닉스 HEVH</h2>
        <h4 style="color:#64748b;">LOSSTIME + SCRAP 분석 대시보드</h4>
    </div>""", unsafe_allow_html=True)
    _,col,_ = st.columns([1,1.2,1])
    with col:
        with st.form("login_form"):
            st.markdown("#### 🔐 로그인")
            u = st.text_input("아이디", placeholder="ID 입력")
            p = st.text_input("비밀번호", type="password", placeholder="PW 입력")
            if st.form_submit_button("로그인", use_container_width=True,
                                     type="primary"):
                users = check_login()
                if u in users and users[u]==p:
                    st.session_state["logged_in"]=True
                    st.session_state["username"]=u
                    st.rerun()
                else:
                    st.error("❌ 아이디 또는 비밀번호가 틀렸습니다.")

# ─────────────────────────────────────────────
# 파싱 엔진
# ─────────────────────────────────────────────
def classify_loss_type(text):
    if not text: return ("ETC","기타")
    t = str(text).lower()
    for code,name,kws in LOSS_TYPE_RULES:
        for kw in kws:
            if kw in t: return (code,name)
    return ("ETC","기타")

def classify_scrap_cause(comment):
    if not comment: return ("AUTO","자동판정(ATE)")
    t = str(comment).lower()
    for code,name,kws in SCRAP_CAUSE_RULES:
        for kw in kws:
            if kw in t: return (code,name)
    return ("ETC","기타")

def normalize_scrap_line(raw):
    s = str(raw or "").strip().upper()
    s = re.sub(r'^PA\s*0*(\d+)', lambda m: f"SA{int(m.group(1)):02d}", s)
    s = re.sub(r'^PS\s*0*(\d+)', lambda m: f"PS{int(m.group(1)):02d}", s)
    s = re.sub(r'^PM\s*0*(\d+)', lambda m: f"MI {int(m.group(1))}", s)
    return s

def extract_slot_causes(cause_row, slots):
    result = {s:"" for s in slots}
    if not cause_row: return result
    cells = list(cause_row)[2:2+len(slots)]
    for idx,cell in enumerate(cells):
        if idx>=len(slots): break
        if cell is not None and str(cell).strip() not in ["","None","—","-"]:
            result[slots[idx]] = " ".join(str(cell).strip().split())
    return result

def extract_model_per_slot(model_row, slots):
    result = {s:"" for s in slots}
    if not model_row: return result
    cells = list(model_row)[2:2+len(slots)]
    last  = ""
    for idx,cell in enumerate(cells):
        if idx>=len(slots): break
        v  = str(cell or "").strip()
        mm = re.search(r'(L\d{2}[A-Z0-9_\-\.]+)', v, re.I)
        if mm:   result[slots[idx]]=mm.group(1); last=mm.group(1)
        elif v and v.lower() not in ["none","—","-",""]:
            result[slots[idx]]=v; last=v
        elif last: result[slots[idx]]=last
    return result

def parse_losstime(val):
    if val is None: return 0.0
    s = str(val).strip()
    if s.startswith("(") and s.endswith(")"): return 0.0
    m = re.search(r'(\d+\.?\d*)', s)
    return max(0.0, float(m.group(1)) if m else 0.0)

def is_line_cell(val):
    return bool(re.match(r'^(SA\s*\d+|PA\s*\d+|PS\s*\d+|MI\s*\d+)',
                         str(val or "").strip(), re.I))

def get_label(row):
    return str(row[1] or "").strip().upper() if len(row)>1 else ""

def detect_process(fn):
    fu = fn.upper()
    if "AI REPORT"  in fu: return "AI"
    if "SMD REPORT" in fu: return "SMT"
    if "MI TIME"    in fu: return "MI"
    if "SCRAP" in fu or "SCRAB" in fu or "스크랩" in fu: return "SCRAP"
    return "UNKNOWN"

def parse_date(text):
    if not text: return None
    t = re.sub(r'\b(DAY|NIGHT)\b','',str(text),flags=re.I).strip()
    m = re.search(r'\b(\d{1,2})\.(\d{1,2})\b', t)
    if m:
        a,b = int(m.group(1)),int(m.group(2))
        if 1<=a<=31 and 1<=b<=12: return f"2026-{b:02d}-{a:02d}"
    m = re.search(r'(\d{1,2})/(\d{1,2})', t)
    if m:
        a,b = int(m.group(1)),int(m.group(2))
        if 1<=a<=31 and 1<=b<=12: return f"2026-{b:02d}-{a:02d}"
    return None

def detect_shift(sn, fn):
    return "NIGHT" if "NIGHT" in (sn+fn).upper() else "DAY"

def normalize_line(raw):
    s = str(raw).strip()
    s = re.sub(r':.*$','',s).strip()
    s = re.sub(r'(?i)^PA\s*0*(\d+)', lambda m: f"SA{int(m.group(1)):02d}", s)
    s = re.sub(r'(?i)^SA\s*0*(\d+)', lambda m: f"SA{int(m.group(1)):02d}", s)
    s = re.sub(r'(?i)^PS\s*0*(\d+)', lambda m: f"PS{int(m.group(1)):02d}", s)
    s = re.sub(r'(?i)^MI\s*0*(\d+)', lambda m: f"MI {m.group(1)}", s)
    return s.strip()

def parse_sheet(ws, process, date_str, shift):
    records = []
    rows    = list(ws.iter_rows(values_only=True))
    slots   = NIGHT_SLOTS if shift=="NIGHT" else DAY_SLOTS
    i = 0
    while i < len(rows):
        row = rows[i]
        c1  = row[1] if len(row)>1 else None
        if not is_line_cell(c1): i+=1; continue
        line = normalize_line(str(c1))
        model_row=loss_row=cause_row=action_row=None
        for j in range(i+1, min(i+14,len(rows))):
            r=rows[j]; lbl=get_label(r)
            if is_line_cell(r[1] if len(r)>1 else None) and j>i+1: break
            if "MODEL"    in lbl and model_row  is None: model_row  = r
            if "LOSSTIME" in lbl and loss_row   is None: loss_row   = r
            elif lbl=="CAUSE"  and cause_row  is None: cause_row  = r
            elif lbl=="ACTION" and action_row is None: action_row = r
        if loss_row:
            loss_vals=[]
            for c in range(2, 2+len(slots)):
                try:
                    v=loss_row[c] if c<len(loss_row) else None
                    s=str(v or "")
                    mm=re.search(r'(\d+)\s*min',s,re.I)
                    loss_vals.append(float(mm.group(1)) if mm else parse_losstime(v))
                except: loss_vals.append(0.0)
            while len(loss_vals)<len(slots): loss_vals.append(0.0)
            loss_vals   = [max(0.0,v) for v in loss_vals]
            total       = sum(loss_vals)
            models      = extract_model_per_slot(model_row, slots)
            slot_causes = extract_slot_causes(cause_row, slots)
            cause_all   = " | ".join(v for v in slot_causes.values() if v)
            unique_c    = set(v for v in slot_causes.values() if v)
            complexity  = "복합" if len(unique_c)>1 else "단일"
            action      = " | ".join(
                str(v) for v in (list(action_row[2:9]) if action_row else [])
                if v and str(v).strip())
            for idx,slot in enumerate(slots):
                lv=loss_vals[idx] if idx<len(loss_vals) else 0.0
                if lv>0:
                    cs=slot_causes.get(slot,"")
                    if not cs:
                        for s2 in slots:
                            if slot_causes.get(s2,""): cs=slot_causes[s2]; break
                    code,name=classify_loss_type(cs)
                    records.append({
                        "date":date_str,"shift":shift,"process":process,
                        "line":line,"time_slot":slot,"model":models.get(slot,""),
                        "loss_min":round(lv,1),"loss_type_code":code,
                        "loss_type_name":name,"complexity":complexity,
                        "loss_detail":cs,"action":action
                    })
            if total>0:
                code_t,name_t=classify_loss_type(cause_all)
                records.append({
                    "date":date_str,"shift":shift,"process":process,
                    "line":line,"time_slot":"TOTAL","model":models.get(slots[0],""),
                    "loss_min":round(total,1),"loss_type_code":code_t,
                    "loss_type_name":name_t,"complexity":complexity,
                    "loss_detail":cause_all,"action":action
                })
        i+=1
    return records

def parse_scrap_file(uploaded_file):
    records=[]
    try:
        fn=uploaded_file.name.lower()
        df=pd.read_excel(uploaded_file,
                         engine="xlrd" if fn.endswith(".xls") else "openpyxl")
    except Exception as e:
        st.error(f"스크랩 파일 읽기 실패: {e}"); return pd.DataFrame()
    df.columns=[str(c).strip() for c in df.columns]
    for r in ["Work Date","Start Line","Tran Comment"]:
        if r not in df.columns:
            st.warning(f"컬럼 없음: {r}"); return pd.DataFrame()
    for _,row in df.iterrows():
        try:
            work_date  =str(row.get("Work Date","")).strip()[:10]
            start_line =normalize_scrap_line(row.get("Start Line",""))
            comment    =str(row.get("Tran Comment","")).strip()
            model_desc =str(row.get("Model Mat Desc","")).strip()
            oper_desc  =str(row.get("Oper Desc","")).strip()
            result_grp =str(row.get("Result Group","")).strip()
            result_cd  =str(row.get("Result Code","")).strip()
            result_desc=str(row.get("Result Desc","")).strip()
            reason_cd  =str(row.get("Reason Code","")).strip()
            reason_desc=str(row.get("Reason Desc","")).strip()
            qty        =int(row.get("Qty",1) or 1)
            mm=re.search(r'(L\d{2}[A-Z0-9_\-\.]+)',model_desc,re.I)
            model  =mm.group(1) if mm else model_desc
            process=OPER_MAP.get(oper_desc,"기타")
            cc,cn  =classify_scrap_cause(comment)
            is_auto="Y" if "ATUO_SCRAP" in comment.upper() or \
                           "AUTO_SCRAP" in comment.upper() else "N"
            records.append({
                "work_date":work_date,"process":process,"line":start_line,
                "model":model,"qty":qty,"cause_code":cc,"cause_name":cn,
                "result_group":result_grp,"result_code":result_cd,
                "result_desc":result_desc,"reason_code":reason_cd,
                "reason_desc":reason_desc,"is_auto":is_auto,"comment":comment,
            })
        except: continue
    return pd.DataFrame(records)

def parse_files(uploaded_files):
    loss_records=[]; scrap_list=[]
    prog=st.progress(0); status=st.empty()
    for fi,uf in enumerate(uploaded_files):
        fn=uf.name; process=detect_process(fn)
        if process=="UNKNOWN":
            status.warning(f"⏭️ 스킵: {fn}")
        elif process=="SCRAP":
            status.info(f"📂 스크랩: {fn}")
            sdf=parse_scrap_file(uf)
            if not sdf.empty:
                scrap_list.append(sdf); status.success(f"✅ {len(sdf):,}건")
        else:
            fd=parse_date(fn) or "UNKNOWN"
            status.info(f"📂 {fn} → {process}/{fd}")
            try: wb=load_workbook(uf,data_only=True)
            except Exception as e:
                status.error(f"열기실패: {e}")
                prog.progress((fi+1)/len(uploaded_files)); continue
            for sn in wb.sheetnames:
                ws=wb[sn]
                if isinstance(ws,Chartsheet): continue
                shift=detect_shift(sn,fn)
                ds=fd if fd!="UNKNOWN" else (parse_date(sn) or "UNKNOWN")
                try: loss_records.extend(parse_sheet(ws,process,ds,shift))
                except Exception as e: st.warning(f"파싱오류[{sn}]: {e}")
        prog.progress((fi+1)/len(uploaded_files))
    status.empty(); prog.empty()
    ldf=pd.DataFrame(loss_records)
    if not ldf.empty and "date" in ldf.columns:
        ldf=ldf[ldf["date"]!="UNKNOWN"].reset_index(drop=True)
    sdf=pd.concat(scrap_list,ignore_index=True) if scrap_list else pd.DataFrame()
    return ldf, sdf

# ─────────────────────────────────────────────
# DB 관리
# ─────────────────────────────────────────────
def load_db():
    df=github_load_csv(DB_PATH)
    if not df.empty and "date" in df.columns:
        df=df[df["date"]!="UNKNOWN"].reset_index(drop=True)
    return df

def save_db(df):       return github_save_csv(df,DB_PATH,"LOSSTIME DB 업데이트")
def load_scrap_db():   return github_load_csv(SCRAP_DB)
def save_scrap_db(df): return github_save_csv(df,SCRAP_DB,"SCRAP DB 업데이트")

def merge_db(existing, new_df):
    if existing.empty: return new_df
    if new_df.empty:   return existing
    combined=pd.concat([existing,new_df],ignore_index=True)
    key=["date","shift","process","line","time_slot"]
    combined=combined.drop_duplicates(
        subset=[c for c in key if c in combined.columns],keep="last")
    return combined.sort_values(
        ["date","shift","process","line","time_slot"]).reset_index(drop=True)

def merge_scrap_db(existing, new_df):
    if existing.empty: return new_df
    if new_df.empty:   return existing
    combined=pd.concat([existing,new_df],ignore_index=True)
    key=["work_date","line","model","comment"]
    combined=combined.drop_duplicates(
        subset=[c for c in key if c in combined.columns],keep="last")
    return combined.sort_values("work_date").reset_index(drop=True)

# ─────────────────────────────────────────────
# 엑셀 다운로드
# ─────────────────────────────────────────────
def to_excel(df):
    buf=io.BytesIO()
    with pd.ExcelWriter(buf,engine="xlsxwriter") as w:
        tdf=df[df["time_slot"]=="TOTAL"] if "time_slot" in df.columns else df
        ls=(tdf.groupby(["process","line"])["loss_min"]
            .sum().reset_index().sort_values("loss_min",ascending=False)
            .rename(columns={"loss_min":"누계손실(분)"}))
        ls["누계손실(분)"]=ls["누계손실(분)"].round(1)
        ls["누계(시간)"]=(ls["누계손실(분)"]/60).round(1)
        ls.insert(0,"순위",range(1,len(ls)+1))
        ls.to_excel(w,sheet_name="라인별누계",index=False)
        ts=(tdf.groupby("loss_type_name")["loss_min"]
            .agg(["sum","count"]).reset_index()
            .sort_values("sum",ascending=False)
            .rename(columns={"loss_type_name":"유형",
                             "sum":"손실(분)","count":"건수"}))
        ts["손실(분)"]=ts["손실(분)"].round(1)
        ts.to_excel(w,sheet_name="손실유형별",index=False)
        df.to_excel(w,sheet_name="원본데이터",index=False)
    return buf.getvalue()

def to_excel_scrap(df):
    buf=io.BytesIO()
    with pd.ExcelWriter(buf,engine="xlsxwriter") as w:
        (df.groupby("cause_name")["qty"].sum().reset_index()
         .sort_values("qty",ascending=False)
         .rename(columns={"cause_name":"원인","qty":"수량"})
         .to_excel(w,sheet_name="원인별",index=False))
        (df.groupby(["process","line"])["qty"].sum().reset_index()
         .sort_values("qty",ascending=False)
         .rename(columns={"qty":"수량"})
         .to_excel(w,sheet_name="라인별",index=False))
        df.to_excel(w,sheet_name="원본데이터",index=False)
    return buf.getvalue()

# ─────────────────────────────────────────────
# 초기화 함수
# ─────────────────────────────────────────────
def reset_filters():
    keys_to_clear = [
        "kpi_focus","date_range_s","date_range_e",
        "proc_pie_chart","type_chart","line_chart",
        "scrap_cause_chart","scrap_line_chart",
    ]
    for k in keys_to_clear:
        if k in st.session_state:
            del st.session_state[k]

# ─────────────────────────────────────────────
# 메인 대시보드
# ─────────────────────────────────────────────
def dashboard():
    # ★ 헤더 클릭 → 초기화
    col_hd, col_reset = st.columns([6,1])
    with col_hd:
        st.markdown("""
        <div class="main-header">
            <h2>🏭 한솔테크닉스 HEVH — LOSSTIME + SCRAP 분석 대시보드</h2>
            <p style="margin:0;opacity:0.85">AI / SMT / PBA(MI) 공정 | 호치민 법인</p>
        </div>""", unsafe_allow_html=True)
    with col_reset:
        st.markdown("<br>", unsafe_allow_html=True)
        if st.button("🏠 홈", use_container_width=True, help="필터 초기화"):
            reset_filters()
            st.rerun()

    # ── 사이드바 ──
    with st.sidebar:
        st.markdown(f"👤 **{st.session_state.get('username','')}** 님")
        if st.button("🚪 로그아웃", use_container_width=True):
            st.session_state["logged_in"]=False; st.rerun()
        st.divider()
        st.header("📂 파일 업로드")
        st.caption("LOSSTIME + SCRAP 동시 가능")
        uploaded=st.file_uploader("xlsx / xls (여러 개)",
                                  type=["xlsx","xls"],accept_multiple_files=True)
        if uploaded:
            if st.button("🚀 분석 시작 / DB 누적",
                         type="primary",use_container_width=True):
                with st.spinner("파싱 중..."):
                    nl,ns=parse_files(uploaded)
                if not nl.empty:
                    ex=load_db(); mg=merge_db(ex,nl)
                    with st.spinner("GitHub 저장..."): ok=save_db(mg)
                    st.session_state["df"]=mg
                    st.success(f"✅ LOSSTIME {len(nl):,}건 → "
                               f"{'GitHub저장' if ok else '로컬저장'}")
                if not ns.empty:
                    es=load_scrap_db(); ms=merge_scrap_db(es,ns)
                    with st.spinner("GitHub 저장..."): ok2=save_scrap_db(ms)
                    st.session_state["scrap_df"]=ms
                    st.success(f"✅ SCRAP {len(ns):,}건 → "
                               f"{'GitHub저장' if ok2 else '로컬저장'}")
                if nl.empty and ns.empty: st.error("파싱 데이터 없음")
        st.divider()
        if st.button("💾 DB 불러오기 (GitHub)",use_container_width=True):
            with st.spinner("로드 중..."):
                db=load_db(); sdb=load_scrap_db()
            if not db.empty:
                st.session_state["df"]=db
                st.success(f"✅ LOSSTIME {len(db):,}건")
            if not sdb.empty:
                st.session_state["scrap_df"]=sdb
                st.success(f"✅ SCRAP {len(sdb):,}건")
            if db.empty and sdb.empty: st.warning("저장된 DB 없음")
        st.divider()
        st.header("🔍 조회 필터")
        df_all=st.session_state.get("df",pd.DataFrame())
        if df_all.empty:
            st.info("데이터를 먼저 불러오세요."); return
        dates=sorted([d for d in df_all["date"].dropna().unique()
                      if d!="UNKNOWN"])
        if not dates:
            st.warning("유효한 날짜 없음"); return
        june=[d for d in dates if d.startswith("2026-06")]
        def_s=june[0]  if june else dates[0]
        def_e=june[-1] if june else dates[-1]
        if len(dates)>=2:
            date_range=st.select_slider("날짜 범위",options=dates,
                                        value=(def_s,def_e))
        else:
            date_range=(dates[0],dates[0]); st.info(f"날짜: {dates[0]}")
        procs=st.multiselect("공정",["AI","SMT","MI"],default=["AI","SMT","MI"])
        shifts=st.multiselect("주야간",["DAY","NIGHT"],default=["DAY","NIGHT"])
        lines_all=sorted(df_all["line"].dropna().unique())
        sel_lines=st.multiselect("라인 (미선택=전체)",lines_all)
        types_all=sorted(df_all["loss_type_name"].dropna().unique())
        sel_types=st.multiselect("손실유형 (미선택=전체)",types_all)
        view_mode=st.radio("조회 단위",["TOTAL(일계)","타임별(A~K)"],horizontal=True)

    # ── 필터 적용 ──
    df=st.session_state.get("df",pd.DataFrame())
    scrap_df=st.session_state.get("scrap_df",pd.DataFrame())
    if df.empty:
        st.info("👈 파일 업로드 또는 DB 불러오기"); return

    mask=(
        (df["date"]>=date_range[0])&(df["date"]<=date_range[1])&
        (df["shift"].isin(shifts or ["DAY","NIGHT"]))&
        (df["process"].isin(procs or ["AI","SMT","MI"]))
    )
    if sel_lines: mask &= df["line"].isin(sel_lines)
    if sel_types: mask &= df["loss_type_name"].isin(sel_types)
    if view_mode=="TOTAL(일계)": mask &= (df["time_slot"]=="TOTAL")
    else:                        mask &= (df["time_slot"]!="TOTAL")

    fdf=df[mask].copy()
    total_df=df[mask&(df["time_slot"]=="TOTAL")].copy()
    if fdf.empty:
        st.warning("⚠️ 조건에 맞는 데이터 없음"); return

    # ── KPI ──
    total_min=round(total_df["loss_min"].sum(),1) if not total_df.empty else 0
    total_hr=round(total_min/60,1)
    n_lines=total_df["line"].nunique() if not total_df.empty else 0
    n_days=fdf["date"].nunique()
    scrap_total=int(scrap_df["qty"].sum()) if not scrap_df.empty else 0

    if "kpi_focus" not in st.session_state:
        st.session_state["kpi_focus"]=None

    c1,c2,c3,c4,c5=st.columns(5)
    for col,val,lbl,key in [
        (c1,f"{total_min:,.1f}분",f"총 손실 ({total_hr}시간)","loss"),
        (c2,f"{n_lines}개","분석 라인 수","line"),
        (c3,f"{n_days}일","분석 일수","date"),
        (c4,f"{scrap_total:,}ea","스크랩 누계","scrap"),
        (c5,f"{len(scrap_df):,}건" if not scrap_df.empty else "0건",
            "스크랩 이력","scrap_hist"),
    ]:
        with col:
            if st.button(f"{val}\n{lbl}",key=f"kpi_{key}",
                         use_container_width=True):
                st.session_state["kpi_focus"]=key

    focus=st.session_state.get("kpi_focus")
    if focus=="loss" and not total_df.empty:
        with st.expander("📊 손실유형별 상세",expanded=True):
            ts=(total_df.groupby("loss_type_name")["loss_min"]
                .sum().reset_index().sort_values("loss_min",ascending=False))
            ts["loss_min"]=ts["loss_min"].round(1)
            st.dataframe(ts.rename(columns={
                "loss_type_name":"유형","loss_min":"손실(분)"}),
                use_container_width=True)
    elif focus=="line" and not total_df.empty:
        with st.expander("📋 라인별 누계 상세",expanded=True):
            ls=(total_df.groupby(["process","line"])["loss_min"]
                .sum().reset_index().sort_values("loss_min",ascending=False))
            ls["loss_min"]=ls["loss_min"].round(1)
            st.dataframe(ls.rename(columns={"loss_min":"손실(분)"}),
                         use_container_width=True)
    elif focus=="date" and not total_df.empty:
        with st.expander("📅 날짜별 상세",expanded=True):
            ds=(total_df.groupby(["date","process"])["loss_min"]
                .sum().reset_index().sort_values("date"))
            ds["loss_min"]=ds["loss_min"].round(1)
            st.dataframe(ds.rename(columns={"loss_min":"손실(분)"}),
                         use_container_width=True)
    elif focus in ["scrap","scrap_hist"] and not scrap_df.empty:
        with st.expander("📛 스크랩 상세",expanded=True):
            st.dataframe(scrap_df,use_container_width=True,height=300)

    st.divider()

    tab1,tab2,tab3,tab4,tab5,tab6,tab7=st.tabs([
        "📊 손실 분석","📈 트렌드","🕐 타임별",
        "📛 스크랩 분석","🔍 상세 조회","🔧 설비보전 PM","⬇️ 다운로드"
    ])

    # ── TAB1 ──
    with tab1:
        if total_df.empty:
            st.warning("TOTAL 데이터 없음")
        else:
            col_l,col_r=st.columns(2)
            with col_l:
                st.subheader("손실유형별 누계 (클릭 → 상세)")
                ts=(total_df.groupby("loss_type_name")["loss_min"]
                    .sum().reset_index()
                    .sort_values("loss_min",ascending=False)
                    .rename(columns={"loss_type_name":"유형","loss_min":"손실(분)"}))
                ts["손실(분)"]=ts["손실(분)"].round(1)
                fig=px.bar(ts,x="손실(분)",y="유형",orientation="h",
                           color="유형",color_discrete_map=TYPE_COLOR,height=520)
                fig.update_layout(showlegend=False,
                                  margin=dict(l=0,r=0,t=20,b=0),
                                  yaxis=dict(categoryorder="total ascending"),
                                  xaxis=dict(rangemode="tozero"))
                ct=st.plotly_chart(fig,use_container_width=True,
                                   on_select="rerun",key="type_chart")
                if ct and ct.get("selection",{}).get("points"):
                    sn2=ct["selection"]["points"][0]
                    sn2=sn2.get("label") or sn2.get("y","")
                    if sn2:
                        st.markdown(f"""
                        <div class="detail-box">
                        <b>📋 [{sn2}] 상세 내역</b>
                        </div>""", unsafe_allow_html=True)
                        dd=fdf[fdf["loss_type_name"]==sn2][
                            ["date","shift","line","time_slot","model",
                             "loss_min","loss_detail"]
                        ].sort_values(["date","line"])
                        dd["loss_min"]=dd["loss_min"].round(1)
                        st.dataframe(dd.reset_index(drop=True),
                                     use_container_width=True,height=300)

            with col_r:
                st.subheader("라인별 누계 TOP 15 (클릭 → 상세)")
                ls=(total_df.groupby(["process","line"])["loss_min"]
                    .sum().reset_index()
                    .sort_values("loss_min",ascending=False).head(15))
                ls["loss_min"]=ls["loss_min"].round(1)
                fig2=px.bar(ls,x="line",y="loss_min",
                            color="process",color_discrete_map=PROC_COLOR,
                            height=520,labels={"loss_min":"손실(분)","line":"라인"})
                fig2.update_layout(margin=dict(l=0,r=0,t=20,b=0),
                                   yaxis=dict(rangemode="tozero"))
                cl=st.plotly_chart(fig2,use_container_width=True,
                                   on_select="rerun",key="line_chart")
                if cl and cl.get("selection",{}).get("points"):
                    sl=cl["selection"]["points"][0].get("x","")
                    if sl:
                        st.markdown(f"""
                        <div class="detail-box">
                        <b>📋 [{sl}] 라인 상세</b>
                        </div>""", unsafe_allow_html=True)
                        ld=fdf[fdf["line"]==sl][
                            ["date","shift","time_slot","model",
                             "loss_min","loss_type_name","loss_detail"]
                        ].sort_values(["date","time_slot"])
                        ld["loss_min"]=ld["loss_min"].round(1)
                        st.dataframe(ld.reset_index(drop=True),
                                     use_container_width=True,height=300)

            # ★ 공정별 비중 파이차트 + 클릭 → 손실유형 바차트
            st.subheader("공정별 비중 (클릭 → 손실유형 확인)")
            ps=total_df.groupby("process")["loss_min"].sum().reset_index()
            ps=ps[ps["process"].isin(["AI","SMT","MI"])]
            ps["loss_min"]=ps["loss_min"].round(1)

            col_pie, col_type = st.columns([1, 1.2])
            with col_pie:
                fig3=px.pie(ps,values="loss_min",names="process",
                            color="process",color_discrete_map=PROC_COLOR,
                            height=380)
                fig3.update_traces(
                    textposition="inside",
                    textinfo="percent+label",
                    hovertemplate="<b>%{label}</b><br>손실: %{value:,.1f}분<br>비중: %{percent}<extra></extra>")
                fig3.update_layout(margin=dict(l=0,r=0,t=20,b=0))
                cp_click=st.plotly_chart(fig3,use_container_width=True,
                                         on_select="rerun",key="proc_pie_chart")

           with col_type:
    sel_proc = ""
    if cp_click and cp_click.get("selection",{}).get("points"):
        pt = cp_click["selection"]["points"][0]
        # ★ 파이차트 클릭 키 전체 탐색
        sel_proc = (pt.get("label") or
                    pt.get("customdata") or
                    pt.get("text") or
                    pt.get("name") or "")
        if isinstance(sel_proc, list):
            sel_proc = sel_proc[0] if sel_proc else ""
        sel_proc = str(sel_proc).strip()
        # AI/SMT/MI 중 하나인지 검증
        if sel_proc not in ["AI","SMT","MI"]:
            sel_proc = ""

    # 클릭 없거나 인식 실패 → 가장 큰 공정 기본값
    if not sel_proc:
        sel_proc = ps.sort_values("loss_min",ascending=False).iloc[0]["process"] \
                   if not ps.empty else "MI"

                if sel_proc:
                    proc_type=(total_df[total_df["process"]==sel_proc]
                               .groupby("loss_type_name")["loss_min"]
                               .sum().reset_index()
                               .sort_values("loss_min",ascending=True)
                               .rename(columns={"loss_type_name":"유형",
                                                "loss_min":"손실(분)"}))
                    proc_type["손실(분)"]=proc_type["손실(분)"].round(1)
                    st.markdown(f"**📊 {sel_proc} 공정 손실유형**")
                    ft=px.bar(proc_type,x="손실(분)",y="유형",
                              orientation="h",
                              color="유형",
                              color_discrete_map=TYPE_COLOR,
                              height=380)
                    ft.update_layout(showlegend=False,
                                     margin=dict(l=0,r=0,t=10,b=0),
                                     xaxis=dict(rangemode="tozero"),
                                     yaxis=dict(categoryorder="total ascending"))
                    st.plotly_chart(ft,use_container_width=True)

    # ── TAB2 ──
    with tab2:
        if total_df.empty:
            st.warning("TOTAL 데이터 없음")
        else:
            st.subheader("날짜별 손실 트렌드")
            dt=(total_df.groupby(["date","process"])["loss_min"].sum().reset_index())
            dt["loss_min"]=dt["loss_min"].round(1)
            fig4=px.line(dt,x="date",y="loss_min",color="process",
                         color_discrete_map=PROC_COLOR,markers=True,height=380,
                         labels={"loss_min":"손실(분)","date":"날짜"})
            fig4.update_layout(margin=dict(l=0,r=0,t=20,b=0),
                               yaxis=dict(rangemode="tozero"))
            st.plotly_chart(fig4,use_container_width=True)

            st.subheader("DAY vs NIGHT")
            sc=(total_df.groupby(["shift","loss_type_name"])["loss_min"].sum().reset_index())
            sc["loss_min"]=sc["loss_min"].round(1)
            fig5=px.bar(sc,x="loss_type_name",y="loss_min",color="shift",
                        color_discrete_map={"DAY":"#f59e0b","NIGHT":"#6366f1"},
                        barmode="group",height=380,
                        labels={"loss_min":"손실(분)","loss_type_name":"손실유형"})
            fig5.update_layout(margin=dict(l=0,r=0,t=20,b=0),
                               xaxis_tickangle=-30,yaxis=dict(rangemode="tozero"))
            st.plotly_chart(fig5,use_container_width=True)

    # ── TAB3 ──
    with tab3:
        st.subheader("🕐 타임별 손실 히트맵")
        slot_df=df[
            (df["date"]>=date_range[0])&(df["date"]<=date_range[1])&
            (df["shift"].isin(shifts or ["DAY","NIGHT"]))&
            (df["process"].isin(procs or ["AI","SMT","MI"]))&
            (df["time_slot"]!="TOTAL")
        ].copy()
        if sel_lines: slot_df=slot_df[slot_df["line"].isin(sel_lines)]
        if slot_df.empty:
            st.warning("타임별 데이터 없음")
        else:
            slot_order=["A","B","C","D","E","F","G","H","I","J","K"]
            pivot=(slot_df.groupby(["line","time_slot"])["loss_min"].sum().reset_index())
            pt=pivot.pivot(index="line",columns="time_slot",
                           values="loss_min").fillna(0).round(1)
            pt=pt[[c for c in slot_order if c in pt.columns]]
            fh=px.imshow(pt,color_continuous_scale="Reds",aspect="auto",
                         height=max(400,len(pt)*30),
                         labels={"x":"시간대","y":"라인","color":"손실(분)"},
                         title="라인 × 시간대 손실 히트맵")
            fh.update_layout(margin=dict(l=0,r=0,t=40,b=0))
            st.plotly_chart(fh,use_container_width=True)

            st.subheader("시간대별 손실 합계")
            ss=(slot_df.groupby(["time_slot","process"])["loss_min"].sum().reset_index())
            ss["loss_min"]=ss["loss_min"].round(1)
            ss["time_slot"]=pd.Categorical(ss["time_slot"],
                                            categories=slot_order,ordered=True)
            ss=ss.sort_values("time_slot")
            fs=px.bar(ss,x="time_slot",y="loss_min",color="process",
                      color_discrete_map=PROC_COLOR,barmode="stack",height=350,
                      labels={"loss_min":"손실(분)","time_slot":"시간대"})
            fs.update_layout(margin=dict(l=0,r=0,t=20,b=0),
                             yaxis=dict(rangemode="tozero"))
            st.plotly_chart(fs,use_container_width=True)

            st.subheader("타임별 상세")
            sd=(slot_df.groupby(["date","shift","line","time_slot",
                                  "loss_type_name","loss_detail"])["loss_min"]
                .sum().reset_index().sort_values(["date","line","time_slot"]))
            sd["loss_min"]=sd["loss_min"].round(1)
            st.dataframe(sd,use_container_width=True,height=400)

    # ── TAB4: 스크랩 ──
    with tab4:
        st.subheader("📛 스크랩 분석")
        if scrap_df.empty:
            st.info("스크랩 데이터 없음. 파일명에 'SCRAP'/'스크랩' 포함 후 업로드.")
        else:
            sd_dates=sorted(scrap_df["work_date"].dropna().unique())
            if len(sd_dates)>=2:
                sr=st.select_slider("스크랩 날짜 범위",options=sd_dates,
                                    value=(sd_dates[0],sd_dates[-1]),key="scrap_date")
                sdf2=scrap_df[(scrap_df["work_date"]>=sr[0])&
                              (scrap_df["work_date"]<=sr[1])].copy()
            else:
                sdf2=scrap_df.copy()
            show_auto=st.checkbox("자동판정(ATE) 포함",value=True)
            if not show_auto: sdf2=sdf2[sdf2["is_auto"]=="N"]
            if sdf2.empty:
                st.warning("조건에 맞는 스크랩 데이터 없음")
            else:
                k1,k2,k3,k4=st.columns(4)
                for col,val,lbl in [
                    (k1,f"{int(sdf2['qty'].sum()):,}ea","총 스크랩"),
                    (k2,f"{len(sdf2):,}건","스크랩 이력"),
                    (k3,f"{sdf2['model'].nunique()}종","모델 수"),
                    (k4,f"{sdf2['line'].nunique()}개","발생 라인"),
                ]:
                    col.markdown(f"""
                    <div class="metric-box">
                        <div class="metric-val">{val}</div>
                        <div class="metric-lbl">{lbl}</div>
                    </div>""", unsafe_allow_html=True)
                st.divider()
                ca,cb=st.columns(2)
                with ca:
                    st.subheader("원인별 파레토 (클릭 → 상세)")
                    cg=(sdf2.groupby("cause_name")["qty"].sum().reset_index()
                        .sort_values("qty",ascending=False))
                    cg["누계%"]=(cg["qty"].cumsum()/cg["qty"].sum()*100).round(1)
                    fp=go.Figure()
                    fp.add_bar(x=cg["cause_name"],y=cg["qty"],name="수량",
                               marker_color="#ef4444")
                    fp.add_scatter(x=cg["cause_name"],y=cg["누계%"],
                                   mode="lines+markers",name="누계%",yaxis="y2",
                                   line=dict(color="#1e3a5f",width=2))
                    fp.update_layout(
                        yaxis2=dict(overlaying="y",side="right",
                                    range=[0,110],ticksuffix="%"),
                        yaxis=dict(rangemode="tozero"),
                        height=400,margin=dict(l=0,r=0,t=20,b=0),
                        xaxis_tickangle=-30)
                    cc2=st.plotly_chart(fp,use_container_width=True,
                                        on_select="rerun",key="scrap_cause_chart")
                    if cc2 and cc2.get("selection",{}).get("points"):
                        sc2=cc2["selection"]["points"][0].get("x","")
                        if sc2:
                            st.markdown(f"**📋 [{sc2}] 상세**")
                            cd=sdf2[sdf2["cause_name"]==sc2][
                                ["work_date","process","line","model","qty","comment"]
                            ].sort_values("work_date")
                            st.dataframe(cd.reset_index(drop=True),
                                         use_container_width=True,height=250)
                with cb:
                    st.subheader("공정별 비중")
                    pg=sdf2.groupby("process")["qty"].sum().reset_index()
                    fpc=px.pie(pg,values="qty",names="process",
                               color="process",color_discrete_map=PROC_COLOR,height=400)
                    fpc.update_layout(margin=dict(l=0,r=0,t=20,b=0))
                    st.plotly_chart(fpc,use_container_width=True)

                st.subheader("라인별 TOP 15 (클릭 → 상세)")
                lg=(sdf2.groupby(["process","line"])["qty"].sum().reset_index()
                    .sort_values("qty",ascending=False).head(15))
                fl=px.bar(lg,x="line",y="qty",color="process",
                          color_discrete_map=PROC_COLOR,height=350,
                          labels={"qty":"수량","line":"라인"})
                fl.update_layout(margin=dict(l=0,r=0,t=20,b=0),
                                 yaxis=dict(rangemode="tozero"))
                csl=st.plotly_chart(fl,use_container_width=True,
                                    on_select="rerun",key="scrap_line_chart")
                if csl and csl.get("selection",{}).get("points"):
                    ssl=csl["selection"]["points"][0].get("x","")
                    if ssl:
                        st.markdown(f"**📋 [{ssl}] 라인 상세**")
                        ld2=sdf2[sdf2["line"]==ssl][
                            ["work_date","model","qty","cause_name","comment"]
                        ].sort_values("work_date")
                        st.dataframe(ld2.reset_index(drop=True),
                                     use_container_width=True,height=250)

                st.subheader("날짜별 트렌드")
                dg=(sdf2.groupby(["work_date","process"])["qty"].sum().reset_index())
                fdt=px.line(dg,x="work_date",y="qty",color="process",
                            color_discrete_map=PROC_COLOR,markers=True,height=320,
                            labels={"qty":"수량","work_date":"날짜"})
                fdt.update_layout(margin=dict(l=0,r=0,t=20,b=0),
                                  yaxis=dict(rangemode="tozero"))
                st.plotly_chart(fdt,use_container_width=True)

                st.subheader("🔧 개선 우선순위")
                imp=(sdf2[sdf2["is_auto"]=="N"]
                     .groupby(["process","line","cause_name"])["qty"]
                     .agg(["sum","count"]).reset_index()
                     .sort_values("sum",ascending=False)
                     .rename(columns={"sum":"수량","count":"발생횟수"}))
                def sg(row):
                    if row["수량"]>=10 or row["발생횟수"]>=5: return "🔴🔴 즉시개선"
                    elif row["수량"]>=5  or row["발생횟수"]>=3: return "🔴 긴급"
                    elif row["수량"]>=3  or row["발생횟수"]>=2: return "🟠 주의"
                    else: return "🟡 모니터링"
                imp["우선순위"]=imp.apply(sg,axis=1)
                for _,row in imp.head(10).iterrows():
                    g=str(row["우선순위"])
                    cls=("pm-p1" if "즉시" in g or "긴급" in g
                         else "pm-p2" if "주의" in g else "pm-p3")
                    st.markdown(f"""
                    <div class="{cls}">
                        <b>{g}</b> &nbsp;|&nbsp;
                        {row['process']} {row['line']} —
                        {row['cause_name']} &nbsp;|&nbsp;
                        <b>{int(row['수량']):,}ea</b> / {int(row['발생횟수'])}회
                    </div>""", unsafe_allow_html=True)

                st.divider()
                st.subheader("스크랩 상세 내역")
                srch=st.text_input("🔍 키워드 (모델/라인/원인/코멘트)")
                sshow=sdf2.copy()
                if srch:
                    ms2=(sshow["model"].astype(str).str.contains(srch,case=False,na=False)|
                         sshow["line"].astype(str).str.contains(srch,case=False,na=False)|
                         sshow["cause_name"].astype(str).str.contains(srch,case=False,na=False)|
                         sshow["comment"].astype(str).str.contains(srch,case=False,na=False))
                    sshow=sshow[ms2]
                dc=["work_date","process","line","model","qty",
                    "cause_name","is_auto","result_desc","comment"]
                dc=[c for c in dc if c in sshow.columns]
                st.caption(f"총 {len(sshow):,}건")
                st.dataframe(sshow[dc].reset_index(drop=True),
                             use_container_width=True,height=400)

    # ── TAB5 ──
    with tab5:
        st.subheader("상세 데이터 조회")
        srch2=st.text_input("🔍 키워드 (라인/원인/모델)")
        sdf3=fdf.copy()
        if srch2:
            m2=(sdf3["line"].astype(str).str.contains(srch2,case=False,na=False)|
                sdf3["loss_detail"].astype(str).str.contains(srch2,case=False,na=False)|
                sdf3["model"].astype(str).str.contains(srch2,case=False,na=False))
            sdf3=sdf3[m2]
        st.caption(f"총 {len(sdf3):,}건")
        dc2=["date","shift","process","line","time_slot",
             "model","loss_min","loss_type_name","complexity","loss_detail"]
        dc2=[c for c in dc2 if c in sdf3.columns]
        if not sdf3.empty:
            sdf3["loss_min"]=sdf3["loss_min"].round(1)
            st.dataframe(sdf3[dc2].reset_index(drop=True),
                         use_container_width=True,height=500)
        else: st.info("검색 결과 없음")

    # ── TAB6 ──
    with tab6:
        st.subheader("🔧 설비보전 PM 우선순위")
        if total_df.empty:
            st.warning("데이터 없음")
        else:
            pm_types=["Printer불량","Axial불량","RH3삽입불량","Mouter불량",
                      "XGZ불량","설비고장(기타)","Coating/Reflow불량",
                      "AOI/S-AOI불량","Wave Solder불량"]
            pm=(total_df[total_df["loss_type_name"].isin(pm_types)]
                .groupby(["process","line","loss_type_name"])["loss_min"]
                .agg(["sum","count"]).reset_index()
                .sort_values("sum",ascending=False)
                .rename(columns={"sum":"누계손실(분)","count":"발생횟수"}))
            pm["누계손실(분)"]=pm["누계손실(분)"].round(1)
            def pmg(row):
                if row["발생횟수"]>=5 or row["누계손실(분)"]>=500: return "🔴🔴 P1 즉시"
                elif row["발생횟수"]>=3 or row["누계손실(분)"]>=200: return "🔴 P1"
                elif row["발생횟수"]>=2 or row["누계손실(분)"]>=100: return "🟠 P2"
                else: return "🟡 P3"
            if pm.empty:
                st.info("설비 불량 데이터 없음")
            else:
                pm["PM등급"]=pm.apply(pmg,axis=1)
                for _,row in pm.iterrows():
                    g=str(row["PM등급"])
                    cls=("pm-p1" if "P1" in g else "pm-p2" if "P2" in g else "pm-p3")
                    st.markdown(f"""
                    <div class="{cls}">
                        <b>{g}</b> &nbsp;|&nbsp;
                        {row['process']} {row['line']} —
                        {row['loss_type_name']} &nbsp;|&nbsp;
                        누계 <b>{row['누계손실(분)']:,.1f}분</b> /
                        {int(row['발생횟수'])}회
                    </div>""", unsafe_allow_html=True)

    # ── TAB7 ──
    with tab7:
        st.subheader("⬇️ 데이터 다운로드")
        ca2,cb2,cc2=st.columns(3)
        with ca2:
            st.markdown("#### 📊 LOSSTIME 엑셀")
            st.download_button("📥 LOSSTIME 다운로드",data=to_excel(fdf),
                               file_name="HEVH_LOSSTIME_리포트.xlsx",
                               mime="application/vnd.openxmlformats-officedocument"
                                    ".spreadsheetml.sheet",use_container_width=True)
        with cb2:
            st.markdown("#### 📛 SCRAP 엑셀")
            if not scrap_df.empty:
                st.download_button("📥 SCRAP 다운로드",data=to_excel_scrap(scrap_df),
                                   file_name="HEVH_SCRAP_리포트.xlsx",
                                   mime="application/vnd.openxmlformats-officedocument"
                                        ".spreadsheetml.sheet",use_container_width=True)
            else: st.info("스크랩 데이터 없음")
        with cc2:
            st.markdown("#### 📋 CSV")
            st.download_button("📥 CSV 다운로드",
                               data=fdf.to_csv(index=False,encoding="utf-8-sig").encode("utf-8-sig"),
                               file_name="HEVH_LOSSTIME_데이터.csv",
                               mime="text/csv",use_container_width=True)
        st.divider()
        st.markdown("#### 🗄️ DB 현황")
        cd1,cd2=st.columns(2)
        with cd1:
            dba=load_db()
            if not dba.empty:
                st.info(f"LOSSTIME: {len(dba):,}건 | "
                        f"{dba['date'].nunique()}일 | {dba['line'].nunique()}개 라인")
                if st.button("🗑️ LOSSTIME DB 초기화"):
                    github_save_csv(pd.DataFrame(),DB_PATH,"DB초기화")
                    st.session_state.pop("df",None)
                    st.success("초기화 완료"); st.rerun()
        with cd2:
            sdba=load_scrap_db()
            if not sdba.empty:
                st.info(f"SCRAP: {len(sdba):,}건 | "
                        f"{int(sdba['qty'].sum())}ea | {sdba['line'].nunique()}개 라인")
                if st.button("🗑️ SCRAP DB 초기화"):
                    github_save_csv(pd.DataFrame(),SCRAP_DB,"SCRAP DB초기화")
                    st.session_state.pop("scrap_df",None)
                    st.success("초기화 완료"); st.rerun()

# ─────────────────────────────────────────────
# 앱 진입점
# ─────────────────────────────────────────────
def main():
    if "logged_in" not in st.session_state:
        st.session_state["logged_in"]=False
    if not st.session_state["logged_in"]: login_page()
    else: dashboard()

if __name__=="__main__":
    main()
