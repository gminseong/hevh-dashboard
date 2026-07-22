"""
[한솔테크닉스 HEVH] LOSSTIME + SCRAP 분석 대시보드 v5.1
실행: python -m streamlit run dashboard.py
변경: v5.1 - AI대기 유형 분리 / 월별 빠른선택 버튼 / 7월 그래프 표시
"""

import re
import io
import base64
import requests
import pandas as pd
import streamlit as st
import plotly.express as px
import plotly.graph_objects as go
from datetime import date, timedelta
from openpyxl import load_workbook
from openpyxl.chartsheet import Chartsheet
from shipment_alert import render_shipment_alert_tab

st.set_page_config(
    page_title="HEVH 대시보드",
    page_icon="🏭",
    layout="wide",
    initial_sidebar_state="expanded"
)

st.markdown("""
<style>
.block-container {
    padding-top: 1.8rem !important;
    padding-bottom: 1rem !important;
}
.main-header {
    padding: 8px 4px 14px;
    margin-bottom: 4px;
    border-bottom: 2px solid #e2e8f0;
}
.main-header h1 {
    margin: 0 0 4px 0;
    font-size: 26px;
    font-weight: 800;
    color: #0f172a;
    line-height: 1.2;
    word-break: break-word;
}
.main-header p {
    margin: 0;
    font-size: 13px;
    color: #64748b;
}
.kpi-card {
    background: #ffffff;
    border: 1px solid #e2e8f0;
    border-radius: 10px;
    padding: 14px 16px;
    text-align: center;
    box-shadow: 0 1px 3px rgba(0,0,0,0.06);
    margin-bottom: 4px;
}
.kpi-val { font-size: 22px; font-weight: 800; color: #0f172a; }
.kpi-lbl { font-size: 11px; color: #64748b; margin-top: 2px; }
.kpi-card-green { background:#f0fdf4; border:1px solid #86efac; border-radius:10px; padding:14px 16px; text-align:center; }
.kpi-card-red   { background:#fef2f2; border:1px solid #fca5a5; border-radius:10px; padding:14px 16px; text-align:center; }
.kpi-val-green  { font-size:22px; font-weight:800; color:#16a34a; }
.kpi-val-red    { font-size:22px; font-weight:800; color:#dc2626; }
.detail-box {
    background: #f8fafc;
    border-left: 4px solid #3b82f6;
    padding: 8px 12px;
    border-radius: 0 6px 6px 0;
    margin: 8px 0;
    font-size: 13px;
}
.pm-p1 {
    background: #fef2f2;
    border-left: 5px solid #dc2626;
    padding: 10px 14px;
    border-radius: 0 8px 8px 0;
    margin: 6px 0;
}
.pm-p2 {
    background: #fff7ed;
    border-left: 5px solid #f97316;
    padding: 10px 14px;
    border-radius: 0 8px 8px 0;
    margin: 6px 0;
}
.pm-p3 {
    background: #fefce8;
    border-left: 5px solid #eab308;
    padding: 10px 14px;
    border-radius: 0 8px 8px 0;
    margin: 6px 0;
}
.section-divider {
    border: none;
    border-top: 1px solid #e2e8f0;
    margin: 20px 0;
}
section[data-testid="stSidebar"] {
    min-width: 240px !important;
    max-width: 300px !important;
    background-color: #f1f5f9 !important;
}
section[data-testid="stSidebar"] .block-container {
    padding: 1rem 0.9rem;
    background-color: #f1f5f9 !important;
}
section[data-testid="stSidebar"] * { color: #1a1a1a !important; }
section[data-testid="stSidebar"] div[data-testid="stButton"] button {
    color: #1e293b !important;
}
section[data-testid="stSidebar"] div[data-testid="stButton"] button[kind="primary"] {
    color: #ffffff !important;
}
section[data-testid="stSidebar"] hr { border-color: #cbd5e1 !important; }
section[data-testid="stSidebar"] [data-testid="stExpander"] {
    background-color: #e8edf2 !important;
    border: 1px solid #cbd5e1 !important;
    border-radius: 8px !important;
}
section[data-testid="stSidebar"] span[data-baseweb="tag"] {
    background-color: #475569 !important;
    color: #ffffff !important;
}
section[data-testid="stSidebar"] span[data-baseweb="tag"] span { color: #ffffff !important; }
section[data-testid="stSidebar"] div[data-testid="stButton"] button {
    background-color: #e2e8f0 !important;
    color: #1e293b !important;
    border: 1px solid #cbd5e1 !important;
    border-radius: 7px !important;
}
section[data-testid="stSidebar"] div[data-testid="stButton"] button:hover {
    background-color: #cbd5e1 !important;
}
section[data-testid="stSidebar"] div[data-testid="stButton"] button[kind="primary"] {
    background-color: #475569 !important;
    color: #ffffff !important;
    border: none !important;
}
section[data-testid="stSidebar"][aria-expanded="false"] {
    min-width: 0px !important;
    max-width: 0px !important;
    overflow: hidden !important;
}
section[data-testid="stMain"] { transition: margin-left 0.3s ease, width 0.3s ease; }
button[kind="header"] { display: none !important; }
[data-testid="baseButton-header"] { display: none !important; }
[data-testid="stStatusWidget"] { display: none !important; }
footer { display: none !important; }
div[data-testid="stButton"] button {
    border-radius: 7px !important;
    font-size: 13px !important;
    font-weight: 500 !important;
}
</style>
""", unsafe_allow_html=True)

# ─────────────────────────────────────────────
# 상수
# ─────────────────────────────────────────────
DAY_SLOTS      = ["A","B","C","D","E","F"]      # AI/SMT DAY: 6 real slots (A-F)
NIGHT_SLOTS    = ["F","G","H","I","J","K"]      # AI/SMT NIGHT: 6 real slots (F-K)
MI_DAY_SLOTS   = ["A","B","C","D","E"]          # MI DAY: only 5 real slots (A-E)
MI_NIGHT_SLOTS = ["F","G","H","I","J"]          # MI NIGHT: only 5 real slots (F-J)
DB_PATH     = "losstime_db.csv"
SCRAP_DB    = "scrap_db.csv"

PROC_COLOR = {
    "AI":  "#ef4444",
    "SMT": "#3b82f6",
    "MI":  "#10b981",
    "기타":"#94a3b8",
}

# ★ v5.1: AI대기 색상 추가
TYPE_COLOR = {
    "모델교체":            "#475569",
    "Magazine부족":        "#64748b",
    "신규OP교육":          "#7c3aed",
    "Printer불량":         "#0369a1",
    "SMT대기":             "#0891b2",
    "AI대기":              "#f97316",   # ★ 신규 추가
    "기타":                "#94a3b8",
    "자재부족":            "#9333ea",
    "AI자재부족":          "#7e22ce",
    "Radial불량":          "#b45309",
    "Mouter불량":          "#0f766e",
    "Axial불량":           "#be185d",
    "Jumper불량":          "#be123c",
    "아일렛/서포트불량":    "#d946ef",
    "Wave Solder불량":     "#1d4ed8",
    "ICT/FT/HiPOT불량":   "#15803d",
    "ATE불량":             "#166534",
    "AOI/S-AOI불량":       "#b45309",
    "SMT Feeder불량":      "#ea580c",
    "계획완료":            "#cbd5e1",
    "설비고장(기타)":      "#dc2626",
    "Silicon/Coating불량": "#047857",
    "Reflow불량":          "#059669",
    "Flux불량":            "#d97706",
    "XGZ불량":             "#7c3aed",
    "SPI불량":             "#075985",
    "Inloader불량":        "#b0bec5",
    "문제없음":            "#e2e8f0",
    "신규모델":            "#6d28d9",
    "NG Buffer불량":       "#ea580c",
    "Camera/Vision불량":   "#0284c7",
    "부품불량":            "#c2410c",
    "RH3삽입불량":         "#92400e",
    "인력부족":            "#64748b",
    "QC검사대기":          "#0284c7",
    "UPH미달":             "#dc2626",
    "샘플/테스트":         "#94a3b8",
}

OPER_MAP = {"P-AUTO":"AI","P-SMD":"SMT","P-PBA":"MI"}

# ─────────────────────────────────────────────
# ★ v5.1 핵심 수정: LOSS_TYPE_RULES
#   - WAITING_AI 신규 추가 (WAITING_SMT 위에 배치)
#   - WAITING_SMT에서 단독 "chờ board" 제거
#   - AI_STOCK은 품질불량 전담 유지
# ─────────────────────────────────────────────
LOSS_TYPE_RULES = [
    ("DONE_PLAN",    "계획완료",           ["done plan","hết plan","kết thúc",
                                            "het plan","ket thuc","no plan"]),

    # ★ 신규: AI대기 — "chờ board AI" 계열 전담 (WAITING_SMT보다 반드시 먼저)
    ("WAITING_AI",   "AI대기",             ["chờ board ai","chờ board từ ai",
                                            "chờ board tu ai","chờ boar ai",
                                            "vừa chờ board ai",
                                            "vừa chạy vừa chờ board ai",
                                            "chờ board ai >> line stop",
                                            "chờ board ai từ",
                                            "board ai từ","waiting board ai",
                                            "stop line chờ board ai",
                                            "cờ board ai","cho board ai",
                                            "chờ board từ ai",
                                            "chờ board ai lỗi",
                                            "chờ ai board",
                                            "đợi board ai",
                                            "thiếu board ai chờ"]),

    # SMT대기 — "chờ board" 단독 키워드 제거 (AI대기와 혼용 방지)
    ("WAITING_SMT",  "SMT대기",            ["waiting semi","waiting pcb","waiting hs",
                                            "wating semi","wating pcb","wating hs",
                                            "waitng","stop line waiting","stop line wating",
                                            "semi pcb","from smd","from smt","from mst",
                                            "clear stock",
                                            "đợi semi","chờ semi",
                                            "wait semi","wait pcb","waiting se mi",
                                            "chờ rad","chờ RAD","hàng chờ",
                                            "chờ board smt","chờ semi smt",
                                            "chờ pcb"]),

    # AI자재부족 — AI 보드 품질불량/수량부족 전담 (대기 아님)
    ("AI_STOCK",     "AI자재부족",         ["ai hết","ai het","thiếu board ai",
                                            "thiếu board","board ai",
                                            "ai stock","vừa chờ board",
                                            "board ai bị","dài chân linh kiện",
                                            "ai scan xót","ai không có board",
                                            "từ ai","đợi ai","chờ ai"]),

    ("MATERIAL",     "자재부족",           ["waiting box","wating box",
                                            "heatsink","heat sink","h/s ",
                                            "packing","mat ","material",
                                            " ic ","ic부족","thiếu ic",
                                            "out of stock",
                                            "thiếu liệu","chờ liệu","đợi liệu",
                                            "thiếu hàng","chờ hàng","thiếu vật tư"]),
    ("MAGAZINE",     "Magazine부족",       ["magazine","magazin","maga sắt","maga sat",
                                            "maga ","mag "]),
    ("NEW_OP",       "신규OP교육",         ["new op","đào tạo","op mới","cover ca","cover work"]),
    ("NEW_MODEL",    "신규모델",           ["lần đầu","model mới","first time",
                                            "new code","setup","set up","new code setup prog",
                                            "sản xuất lần đầu"]),
    ("MANPOWER",     "인력부족",           ["nhân lực","nhan luc","không đủ nhân",
                                            "thiếu người","thiếu cn"]),
    ("FEEDER",       "SMT Feeder불량",     ["feeder","feed error","feed lỗi"]),
    ("MOUTER",       "Mouter불량",         ["moutor","mouter","mounter","mouotr",
                                            "mount ","stopper","băng tải",
                                            "không ổn định","văng linh kiện",
                                            "pickup","pick up","pick-up",
                                            "trục z","mất khí","head ",
                                            "treo máy","kẹt board"]),
    ("PRINTER",      "Printer불량",        ["printer","priter","lỗi keo","tràn keo"]),
    ("SPI",          "SPI불량",            ["spi"]),
    ("INLOADER",     "Inloader불량",       ["inloader","inloarder","unloder","unloader"]),
    ("INSERT_AXIAL", "Axial불량",          ["av131","av 131","av insert","av lỗi",
                                            "axial","ax불","av "]),
    ("INSERT_JUMP",  "Jumper불량",         ["jv13","jv ","jv불","jumper","jump",
                                            "đứt jump","dut jump","cong jump",
                                            "lỗi jump"]),
    ("INSERT_RAD",   "Radial불량",         ["radial","rad ","rad불","rad lỗi","rad rớt",
                                            "rad mất","rad insert",
                                            "rg131","rg 131","rg ","rg불","rg cắm",
                                            "rh3","rh 3","rh5","rh 5",
                                            "rhu","rh ","rhu lỗi",
                                            "pin insert"]),
    ("EYELET",       "아일렛/서포트불량",   ["eyelet","eylet","eye.sp","eye sp",
                                            "eys ","eye ","ey ","ey불",
                                            "sp ","sp불","sp lỗi","support",
                                            "bể eye","máy ey"]),
    ("INSERT_XGZ",   "XGZ불량",            ["xzg","xgz"]),
    ("CHANGE_MODEL", "모델교체",           ["change model","change cod","change ",
                                            "đổi model","tách lot","tách lót","line change",
                                            "làm program","program new","program ver"]),
    ("NG_BUFFER",    "NG Buffer불량",      ["ng buffer","ng buffet","kẹt boar ng",
                                            "buffer ng","kẹt bo"]),
    ("ATE",          "ATE불량",            ["ate error","ate run","ate lỗi",
                                            "ate 1st","ate 1 floor","ate running",
                                            "running 1 floor","running 1floor",
                                            "ate setup","set up ate","setup ate",
                                            "ate no scan","ate stop","ate "]),
    ("ICT_FT",       "ICT/FT/HiPOT불량",  ["ict error","ict no scan","ict lỗi",
                                            "ft error","ft no scan","ft lỗi",
                                            "hipot","hi-pot","hi pot",
                                            "hipot error","cleaning ict",
                                            "ict ","ft ","ict&"]),
    ("AOI_SAOI",     "AOI/S-AOI불량",      ["saoi","s-aoi","aoi error","aoi no scan",
                                            "aoi lỗi","aoi","scan sót","scan nhầm"]),
    ("CAMERA",       "Camera/Vision불량",  ["camera","vision","lỗi camera",
                                            "don't connect"]),
    ("QC_INSPECT",   "QC검사대기",         ["qc kiểm","qc check","kiểm tra board",
                                            "vender","vendor"]),
    ("UPH",          "UPH미달",            ["uph","không đáp ứng"]),
    ("WAVE",         "Wave Solder불량",    ["wave solder","wave"]),
    ("FLUX",         "Flux불량",           ["flux machine","flux error","flux"]),
    ("SILCOAT",      "Silicon/Coating불량",["silicon","coating","nozzle","coatin"]),
    ("REFLOW",       "Reflow불량",         ["reflow"]),
    ("COMP_ERR",     "부품불량",           ["lệch linh kiện","linh kiện lỗi",
                                            "thiếu linh kiện","component",
                                            "lỗi lệch","chân dài","dài chân"]),
    ("EQUIP_FAIL",   "설비고장(기타)",     ["hư","lỗi máy","spare","power off",
                                            "vaccum","vacuum","rail bị",
                                            "stop line xử","stop line xu",
                                            "kẹt dao","ket dao","cutter"]),
    ("SAMPLE",       "샘플/테스트",        ["sample","running sample","kiểm bo"]),
    ("NO_PROBLEM",   "문제없음",           ["no problem","no probplem","no proplem",
                                            "no probem","no proble","no prob",
                                            "không vấn đề"]),
    ("ETC",          "기타",               []),
]

SCRAP_CAUSE_RULES = [
    ("COMP",   "부품불량",    ["component","linh kiện","missing","thiếu"]),
    ("SOLDER", "납불량",      ["solder","hàn","weld"]),
    ("MACH",   "설비불량",    ["machine","máy","equipment"]),
    ("OP",     "작업자불량",  ["operator","công nhân","op lỗi"]),
    ("AUTO",   "자동판정(ATE)",["atuo_scrap","auto_scrap","ate"]),
    ("ETC",    "기타",        []),
]

# ─────────────────────────────────────────────
# GitHub
# ─────────────────────────────────────────────
def get_github_config():
    try: return st.secrets["github"]["token"], st.secrets["github"]["repo"]
    except: return None, None

def github_load_csv(filename):
    token, repo = get_github_config()
    if not token:
        try: return pd.read_csv(filename, encoding="utf-8-sig")
        except FileNotFoundError: return pd.DataFrame()
    url = f"https://api.github.com/repos/{repo}/contents/{filename}"
    headers = {"Authorization":f"token {token}",
               "Accept":"application/vnd.github.v3+json"}
    try:
        r = requests.get(url, headers=headers, timeout=10)
        if r.status_code == 200:
            return pd.read_csv(io.StringIO(
                base64.b64decode(r.json()["content"]).decode("utf-8")))
        return pd.DataFrame()
    except: return pd.DataFrame()

def github_save_csv(df, filename, msg=None):
    token, repo = get_github_config()
    if not token:
        df.to_csv(filename, index=False, encoding="utf-8-sig"); return True
    url = f"https://api.github.com/repos/{repo}/contents/{filename}"
    headers = {"Authorization":f"token {token}",
               "Accept":"application/vnd.github.v3+json"}
    sha = None
    try:
        r = requests.get(url, headers=headers, timeout=10)
        if r.status_code == 200: sha = r.json().get("sha")
    except: pass
    csv_str = df.to_csv(index=False, encoding="utf-8-sig")
    content  = base64.b64encode(csv_str.encode("utf-8-sig")).decode()
    payload  = {"message": msg or f"DB:{filename}", "content": content}
    if sha: payload["sha"] = sha
    try:
        r = requests.put(url, headers=headers, json=payload, timeout=15)
        return r.status_code in [200, 201]
    except: return False

def github_load_xlsx(filename):
    token, repo = get_github_config()
    if not token:
        try: return pd.read_excel(filename)
        except FileNotFoundError: return pd.DataFrame()
    url = f"https://api.github.com/repos/{repo}/contents/{filename}"
    headers = {"Authorization": f"token {token}",
               "Accept": "application/vnd.github.v3+json"}
    try:
        r = requests.get(url, headers=headers, timeout=10)
        if r.status_code == 200:
            content = base64.b64decode(r.json()["content"])
            return pd.read_excel(io.BytesIO(content))
        return pd.DataFrame()
    except: return pd.DataFrame()

def github_save_xlsx(df_or_bytes, filename, msg=None):
    token, repo = get_github_config()
    if not token: return False
    url = f"https://api.github.com/repos/{repo}/contents/{filename}"
    headers = {"Authorization": f"token {token}",
               "Accept": "application/vnd.github.v3+json"}
    if isinstance(df_or_bytes, bytes):
        content = base64.b64encode(df_or_bytes).decode()
    else:
        buf = io.BytesIO()
        df_or_bytes.to_excel(buf, index=False)
        content = base64.b64encode(buf.getvalue()).decode()
    sha = None
    try:
        r = requests.get(url, headers=headers, timeout=10)
        if r.status_code == 200: sha = r.json().get("sha")
    except: pass
    payload = {"message": msg or f"DB:{filename}", "content": content}
    if sha: payload["sha"] = sha
    try:
        r = requests.put(url, headers=headers, json=payload, timeout=15)
        return r.status_code in [200, 201]
    except: return False

# ─────────────────────────────────────────────
# DB
# ─────────────────────────────────────────────
def load_db():
    df = github_load_csv(DB_PATH)
    if not df.empty and "date" in df.columns:
        df = df[df["date"] != "UNKNOWN"].reset_index(drop=True)
    for col in ["target","actual","target_mi","actual_mi","target_ate","actual_ate"]:
        if col not in df.columns: df[col] = 0
    return df

def save_db(df):       return github_save_csv(df, DB_PATH, "LOSSTIME DB")
def load_scrap_db():   return github_load_csv(SCRAP_DB)
def save_scrap_db(df): return github_save_csv(df, SCRAP_DB, "SCRAP DB")

def merge_db(existing, new_df):
    if existing.empty: return new_df
    if new_df.empty:   return existing
    combined = pd.concat([existing, new_df], ignore_index=True)
    key = ["date", "shift", "process", "line", "time_slot", "sub_idx"]
    combined = combined.drop_duplicates(
        subset=[c for c in key if c in combined.columns], keep="last")
    return combined.sort_values(
        ["date","shift","process","line","time_slot"]).reset_index(drop=True)

def merge_scrap_db(existing, new_df):
    if existing.empty: return new_df
    if new_df.empty:   return existing
    combined = pd.concat([existing, new_df], ignore_index=True)
    key = ["work_date","line","model","comment"]
    combined = combined.drop_duplicates(
        subset=[c for c in key if c in combined.columns], keep="last")
    return combined.sort_values("work_date").reset_index(drop=True)

# ─────────────────────────────────────────────
# 파싱 엔진
# ─────────────────────────────────────────────
def classify_loss_type(text):
    if not text: return ("ETC","기타")
    t = str(text).lower()
    for code, name, kws in LOSS_TYPE_RULES:
        for kw in kws:
            if kw in t: return (code, name)
    return ("ETC","기타")

def parse_time_from_text(text):
    t = str(text).strip()
    m = re.search(r'(\d+)\s*min', t, re.I)
    if m: return float(m.group(1))
    m = re.search(r'(\d+)\s*분', t)
    if m: return float(m.group(1))
    m = re.search(r"(\d+)\s*'", t)
    if m: return float(m.group(1))
    m = re.search(r'(\d{1,2}):(\d{2})\s*[-~]\s*(\d{1,2}):(\d{2})', t)
    if m:
        start = int(m.group(1)) * 60 + int(m.group(2))
        end   = int(m.group(3)) * 60 + int(m.group(4))
        if end < start: end += 24 * 60
        diff = end - start
        if 0 < diff < 720: return float(diff)
    return None

def split_loss_detail(loss_detail, total_min):
    if not loss_detail or str(loss_detail).strip() in ["", "None", "–", "-"]:
        return [{"detail": "", "min": total_min, "type_code": "ETC",
                 "type_name": "기타", "sub_idx": 1}]
    raw   = str(loss_detail).strip()
    parts = [p.strip() for p in raw.split("|")]
    parts = [p for p in parts if p]
    filtered = [p for p in parts if classify_loss_type(p)[1] not in ("문제없음","계획완료")]
    if not filtered:
        return []
    seen = {}
    for p in filtered:
        code, name = classify_loss_type(p)
        if code not in seen:
            seen[code] = {"detail": p, "type_code": code, "type_name": name}
        else:
            seen[code]["detail"] += " | " + p
    unique_types = list(seen.values())
    n = len(unique_types)
    per_min = round(total_min / n, 1)
    return [
        {"detail": v["detail"], "min": per_min,
         "type_code": v["type_code"], "type_name": v["type_name"],
         "sub_idx": idx + 1}
        for idx, v in enumerate(unique_types)
    ]

def classify_scrap_cause(comment):
    if not comment: return ("AUTO","자동판정(ATE)")
    t = str(comment).lower()
    for code, name, kws in SCRAP_CAUSE_RULES:
        for kw in kws:
            if kw in t: return (code, name)
    return ("ETC","기타")

def normalize_scrap_line(raw):
    s = str(raw or "").strip().upper()
    s = re.sub(r'^PA\s*0*(\d+)', lambda m: f"SA{int(m.group(1)):02d}", s)
    s = re.sub(r'^PS\s*0*(\d+)', lambda m: f"PS{int(m.group(1)):02d}", s)
    s = re.sub(r'^PM\s*0*(\d+)', lambda m: f"MI {int(m.group(1))}", s)
    return s

def get_label(row):
    return str(row[1] or "").strip().upper() if len(row) > 1 else ""

def detect_process(fn):
    fu = fn.upper()
    if "AI REPORT"  in fu: return "AI"
    if "SMD REPORT" in fu: return "SMT"
    if "MI TIME"    in fu: return "MI"
    if "SCRAP" in fu or "SCRAB" in fu or "스크랩" in fu: return "SCRAP"
    return "UNKNOWN"

def parse_date(text):
    if not text: return None
    import datetime
    if isinstance(text, (datetime.datetime, datetime.date)):
        return text.strftime("%Y-%m-%d")
    t = re.sub(r'\b(DAY|NIGHT|NHU|SMD|AI|MI|REPORT|DAILY|TIME)\b', '', str(text), flags=re.I)
    t = re.sub(r'\s+\d{2}:\d{2}:\d{2}.*$', '', t).strip()
    m = re.search(r'OT\s*(\d{1,2})\.(\d{1,2})', t, re.I)
    if m:
        day, month = int(m.group(1)), int(m.group(2))
        if 1 <= day <= 31 and 1 <= month <= 12:
            return f"2026-{month:02d}-{day:02d}"
    m = re.search(r'\b(\d{1,2})\.(\d{1,2})\b', t)
    if m:
        a, b = int(m.group(1)), int(m.group(2))
        if 1 <= a <= 31 and 1 <= b <= 12:
            return f"2026-{b:02d}-{a:02d}"
    m = re.search(r'(\d{1,2})/(\d{1,2})', t)
    if m:
        a, b = int(m.group(1)), int(m.group(2))
        if 1 <= a <= 31 and 1 <= b <= 12:
            return f"2026-{b:02d}-{a:02d}"
    return None

def detect_shift(sn, fn):
    combined = (sn + fn).upper()
    if "NIGHT" in combined or "NIGH" in combined:
        return "NIGHT"
    if re.search(r'N[Ii][Gg][Hh]', sn + fn):
        return "NIGHT"
    return "DAY"

def find_date_in_sheet(ws):
    rows = list(ws.iter_rows(values_only=True, max_row=20))
    for row in rows:
        for cell in row:
            if cell is None: continue
            s = str(cell)
            m = re.search(r'\((\d{1,2})/(\d{1,2})\)', s)
            if m:
                a, b = int(m.group(1)), int(m.group(2))
                if 1 <= a <= 31 and 1 <= b <= 12:
                    return f"2026-{b:02d}-{a:02d}"
            m2 = re.search(r'\b(\d{1,2})/(\d{2})\b', s)
            if m2:
                a2, b2 = int(m2.group(1)), int(m2.group(2))
                if 1 <= a2 <= 31 and 1 <= b2 <= 12:
                    return f"2026-{b2:02d}-{a2:02d}"
            d = parse_date(s)
            if d: return d
    return None

def sheet_title_date(ws):
    """★ 시트 자체의 'PRODUCTION REPORT BY TIME ...' 제목 셀에서만 날짜를 추출한다.
       주간 마감 직후 업로드된 파일은 NIGHT 시트가 아직 전날 데이터로 남아있는
       경우가 있어(야간 마감 전), 파일명 날짜만 믿으면 잘못된 날짜로 섞여 들어간다.
       제목 셀(REPORT/PRODUCTION 키워드 포함)에서만 날짜를 찾으므로 UPH, 비율 등
       임의의 숫자 셀을 날짜로 오인하는 일은 없다."""
    rows = list(ws.iter_rows(values_only=True, max_row=6))
    for row in rows:
        for cell in row:
            if cell is None: continue
            s = str(cell)
            if not re.search(r'REPORT|PRODUCTION', s, re.I):
                continue
            d = parse_date(s)
            if d: return d
    return None

def normalize_line(raw):
    s = str(raw).strip()
    s = re.sub(r':.*$', '', s).strip()
    # ★ 코드 패턴(SA/PA/PS/MI + 번호 + 선택적 단일 접미문자)만 뽑아내고
    #   뒤에 붙은 부가설명("PS12 Plan line 13")이나 "+PS07" 같은 병기는 버린다.
    #   (안 그러면 "PS12 Plan line 13", "PS06+PS07" 같은 값이 그대로 별도
    #   라인으로 취급되어 히트맵/라인목록에 유령 라인이 생겼음)
    m = re.match(r'(?i)^(SA|PA|PS|MI)\s*0*(\d+)([A-Z]?)(?![a-zA-Z])', s)
    if m:
        prefix, num, suffix = m.group(1).upper(), int(m.group(2)), m.group(3).upper()
        if prefix in ("PA", "SA"):
            return f"SA{num:02d}{suffix}"
        if prefix == "PS":
            return f"PS{num:02d}{suffix}"
        if prefix == "MI":
            return f"MI {num}{suffix}"
    return s.strip()

def is_line_cell(val):
    s = str(val or "").strip()
    if not s: return False
    return bool(re.match(
        r'^(SA\s*\d+|PA\s*\d+|PS\s*\d+|MI\s*\d+)', s, re.I))

def parse_losstime(val):
    if val is None: return 0.0
    s = str(val).strip()
    if s.startswith("(") and s.endswith(")"): return 0.0
    m = re.search(r'(\d+\.?\d*)', s)
    return max(0.0, float(m.group(1)) if m else 0.0)

def extract_slot_causes(cause_row, slots):
    result = {s: "" for s in slots}
    if not cause_row: return result
    cells = list(cause_row)[2:2+len(slots)]
    for idx, cell in enumerate(cells):
        if idx >= len(slots): break
        if cell is not None and str(cell).strip() not in ["", "None", "-"]:
            result[slots[idx]] = " ".join(str(cell).strip().split())
    return result

def extract_model_per_slot(model_row, slots):
    result = {s: "" for s in slots}
    if not model_row: return result
    cells = list(model_row)[2:2+len(slots)]
    last  = ""
    for idx, cell in enumerate(cells):
        if idx >= len(slots): break
        v  = str(cell or "").strip()
        mm = re.search(r'(L\d{2}[A-Z0-9_\-\.]+)', v, re.I)
        if mm: result[slots[idx]] = mm.group(1); last = mm.group(1)
        elif v and v.lower() not in ["none", "-", ""]:
            result[slots[idx]] = v; last = v
        elif last: result[slots[idx]] = last
    return result

def parse_sheet(ws, process, date_str, shift):
    records = []
    rows    = list(ws.iter_rows(values_only=True))
    if process == "MI":
        slots = MI_NIGHT_SLOTS if shift == "NIGHT" else MI_DAY_SLOTS
    else:
        slots = NIGHT_SLOTS if shift == "NIGHT" else DAY_SLOTS
    i = 0
    while i < len(rows):
        row = rows[i]; c1 = row[1] if len(row) > 1 else None
        if not is_line_cell(c1): i += 1; continue

        line = normalize_line(str(c1))
        model_row = loss_row = cause_row = action_row = None
        target_row = actual_row = None

        for j in range(i+1, min(i+16, len(rows))):
            r   = rows[j]; lbl = get_label(r)
            if is_line_cell(r[1] if len(r) > 1 else None) and j > i+1: break
            if "MODEL"    in lbl and model_row  is None: model_row  = r
            if "LOSSTIME" in lbl and loss_row   is None: loss_row   = r
            elif lbl == "CAUSE"  and cause_row  is None: cause_row  = r
            elif lbl == "ACTION" and action_row is None: action_row = r
            elif "TARGET" in lbl and target_row is None: target_row = r
            elif "ACTUAL" in lbl and actual_row is None: actual_row = r

        if loss_row:
            loss_vals = []
            for c in range(2, len(loss_row)):
                v = loss_row[c]
                if v is None or str(v).strip() == "":
                    loss_vals.append(0.0)
                elif isinstance(v, (int, float)):
                    loss_vals.append(max(0.0, float(v)))
                else:
                    s  = str(v).strip()
                    mm = re.search(r'(\d+)\s*min', s, re.I)
                    if mm:
                        loss_vals.append(float(mm.group(1)))
                    else:
                        pv = parse_losstime(v)
                        if pv > 0: loss_vals.append(pv)
                        else: break
            while len(loss_vals) < len(slots): loss_vals.append(0.0)
            loss_vals = loss_vals[:len(slots)]
            loss_vals = [max(0.0, v) for v in loss_vals]

            if process == "MI":
                loss_vals = []
                for s_idx in range(len(slots)):
                    col = 2 + s_idx * 3
                    v   = loss_row[col] if col < len(loss_row) else None
                    loss_vals.append(max(0.0, parse_losstime(v)))
                total = sum(loss_vals)
            else:
                total = sum(loss_vals)

            models      = extract_model_per_slot(model_row, slots)
            slot_causes = extract_slot_causes(cause_row, slots)
            cause_all   = " | ".join(v for v in slot_causes.values() if v)
            if action_row:
                action = " | ".join(
                    str(action_row[2+idx] or "").strip()
                    for idx in range(len(slots))
                    if action_row[2+idx] and str(action_row[2+idx]).strip()
                )
            else:
                action = ""

            unique_c   = set(v for v in slot_causes.values() if v)
            complexity = "복합" if len(unique_c) > 1 else "단일"

            # TARGET / ACTUAL
            if process == "MI":
                mi_target = mi_actual = ate_target = ate_actual = 0.0
                for s_idx in range(len(slots)):
                    col_mi  = 2 + s_idx * 3
                    col_ate = 2 + s_idx * 3 + 2
                    if target_row:
                        try: mi_target  += parse_losstime(target_row[col_mi]  if col_mi  < len(target_row) else None)
                        except: pass
                        try: ate_target += parse_losstime(target_row[col_ate] if col_ate < len(target_row) else None)
                        except: pass
                    if actual_row:
                        try: mi_actual  += parse_losstime(actual_row[col_mi]  if col_mi  < len(actual_row) else None)
                        except: pass
                        try: ate_actual += parse_losstime(actual_row[col_ate] if col_ate < len(actual_row) else None)
                        except: pass
                target_tot = round(mi_target, 0)
                actual_tot = round(mi_actual, 0)
                target_mi  = round(mi_target, 0)
                actual_mi  = round(mi_actual, 0)
                target_ate = round(ate_target, 0)
                actual_ate = round(ate_actual, 0)
            else:
                target_tot = actual_tot = 0.0
                for s_idx in range(len(slots)):
                    col = 2 + s_idx
                    if target_row:
                        try: target_tot += parse_losstime(target_row[col] if col < len(target_row) else None)
                        except: pass
                    if actual_row:
                        try: actual_tot += parse_losstime(actual_row[col] if col < len(actual_row) else None)
                        except: pass
                target_mi = target_ate = actual_mi = actual_ate = 0.0

            if total > 0:
                cause_totals = {}
                for idx, slot in enumerate(slots):
                    lv = loss_vals[idx] if idx < len(loss_vals) else 0.0
                    if lv > 0:
                        cs = slot_causes.get(slot, "") or cause_all
                        code, name = classify_loss_type(cs)
                        if name not in cause_totals:
                            cause_totals[name] = {"min": 0.0, "detail": cs,
                                                   "type_code": code, "type_name": name}
                        cause_totals[name]["min"] += lv
                sub_details = [
                    {"detail": v["detail"], "min": round(v["min"], 1),
                     "type_code": v["type_code"], "type_name": v["type_name"],
                     "sub_idx": i2 + 1}
                    for i2, v in enumerate(cause_totals.values())
                ]
                if not sub_details:
                    sub_details = split_loss_detail(cause_all, total)
                if not sub_details:
                    i += 1; continue

                # TOTAL 행
                if len(sub_details) == 1:
                    sd = sub_details[0]
                    records.append({
                        "date": date_str, "shift": shift, "process": process,
                        "line": line, "time_slot": "TOTAL",
                        "model": models.get(slots[0], ""),
                        "loss_min": round(total, 1),
                        "loss_type_code": sd["type_code"],
                        "loss_type_name": sd["type_name"],
                        "complexity": "단일",
                        "loss_detail": sd["detail"],
                        "sub_idx": 1, "action": action,
                        "target": round(target_tot, 0), "actual": round(actual_tot, 0),
                        "target_mi": round(target_mi, 0), "actual_mi": round(actual_mi, 0),
                        "target_ate": round(target_ate, 0), "actual_ate": round(actual_ate, 0),
                    })
                else:
                    for sd in sub_details:
                        records.append({
                            "date": date_str, "shift": shift, "process": process,
                            "line": line, "time_slot": "TOTAL",
                            "model": models.get(slots[0], ""),
                            "loss_min": sd["min"],
                            "loss_type_code": sd["type_code"],
                            "loss_type_name": sd["type_name"],
                            "complexity": "복합",
                            "loss_detail": sd["detail"],
                            "sub_idx": sd["sub_idx"], "action": action,
                            "target": round(target_tot, 0), "actual": round(actual_tot, 0),
                            "target_mi": round(target_mi, 0), "actual_mi": round(actual_mi, 0),
                            "target_ate": round(target_ate, 0), "actual_ate": round(actual_ate, 0),
                        })

                # 타임별 행
                for idx, slot in enumerate(slots):
                    lv = loss_vals[idx] if idx < len(loss_vals) else 0.0
                    if lv > 0:
                        cs = slot_causes.get(slot, "")
                        if not cs:
                            for s2 in slots:
                                if slot_causes.get(s2, ""): cs = slot_causes[s2]; break
                        code, name = classify_loss_type(cs)
                        records.append({
                            "date": date_str, "shift": shift, "process": process,
                            "line": line, "time_slot": slot,
                            "model": models.get(slot, ""),
                            "loss_min": round(lv, 1),
                            "loss_type_code": code, "loss_type_name": name,
                            "complexity": complexity, "loss_detail": cs,
                            "sub_idx": 1, "action": action,
                            "target": 0, "actual": 0,
                            "target_mi": 0, "actual_mi": 0,
                            "target_ate": 0, "actual_ate": 0,
                        })
            else:
                # ★ 로스타임이 0(계획 100% 달성)이어도 해당 라인의 target/actual은
                #   집계에 반드시 포함되어야 하므로, 손실 상세 없이 TOTAL 1건만 남김
                records.append({
                    "date": date_str, "shift": shift, "process": process,
                    "line": line, "time_slot": "TOTAL",
                    "model": models.get(slots[0], ""),
                    "loss_min": 0.0,
                    "loss_type_code": "NO_PROBLEM", "loss_type_name": "문제없음",
                    "complexity": "단일", "loss_detail": "",
                    "sub_idx": 1, "action": action,
                    "target": round(target_tot, 0), "actual": round(actual_tot, 0),
                    "target_mi": round(target_mi, 0), "actual_mi": round(actual_mi, 0),
                    "target_ate": round(target_ate, 0), "actual_ate": round(actual_ate, 0),
                })
        i += 1
    return records

def parse_scrap_file(uploaded_file):
    records = []
    try:
        fn = uploaded_file.name.lower()
        df = pd.read_excel(uploaded_file,
                           engine="xlrd" if fn.endswith(".xls") else "openpyxl")
    except Exception as e:
        st.error(f"스크랩 읽기 실패: {e}"); return pd.DataFrame()
    df.columns = [str(c).strip() for c in df.columns]
    for r in ["Work Date","Start Line","Tran Comment"]:
        if r not in df.columns:
            st.warning(f"컬럼 없음: {r}"); return pd.DataFrame()
    for _, row in df.iterrows():
        try:
            work_date   = str(row.get("Work Date","")).strip()[:10]
            start_line  = normalize_scrap_line(row.get("Start Line",""))
            comment     = str(row.get("Tran Comment","")).strip()
            model_desc  = str(row.get("Model Mat Desc","")).strip()
            oper_desc   = str(row.get("Oper Desc","")).strip()
            result_grp  = str(row.get("Result Group","")).strip()
            result_cd   = str(row.get("Result Code","")).strip()
            result_desc = str(row.get("Result Desc","")).strip()
            reason_cd   = str(row.get("Reason Code","")).strip()
            reason_desc = str(row.get("Reason Desc","")).strip()
            qty         = int(row.get("Qty", 1) or 1)
            mm          = re.search(r'(L\d{2}[A-Z0-9_\-\.]+)', model_desc, re.I)
            model       = mm.group(1) if mm else model_desc
            process     = OPER_MAP.get(oper_desc, "기타")
            cc, cn      = classify_scrap_cause(comment)
            is_auto     = "Y" if "ATUO_SCRAP" in comment.upper() or \
                                 "AUTO_SCRAP" in comment.upper() else "N"
            records.append({
                "work_date": work_date, "process": process, "line": start_line,
                "model": model, "qty": qty, "cause_code": cc, "cause_name": cn,
                "result_group": result_grp, "result_code": result_cd,
                "result_desc": result_desc, "reason_code": reason_cd,
                "reason_desc": reason_desc, "is_auto": is_auto, "comment": comment
            })
        except: continue
    return pd.DataFrame(records)

def parse_files(uploaded_files):
    loss_records = []; scrap_list = []
    prog = st.progress(0); status = st.empty()
    parsed_keys = set()

    for fi, uf in enumerate(uploaded_files):
        fn = uf.name; process = detect_process(fn)
        if process == "UNKNOWN":
            status.warning(f"skip: {fn}")
        elif process == "SCRAP":
            status.info(f"scrap: {fn}")
            sdf = parse_scrap_file(uf)
            if not sdf.empty:
                scrap_list.append(sdf); status.success(f"OK {len(sdf):,}")
        else:
            fd = parse_date(fn) or "UNKNOWN"
            status.info(f"{fn} -> {process}/{fd}")
            try: wb = load_workbook(uf, data_only=True)
            except Exception as e:
                status.error(f"열기실패: {e}")
                prog.progress((fi+1)/len(uploaded_files)); continue
            for sn in wb.sheetnames:
                ws = wb[sn]
                if isinstance(ws, Chartsheet): continue
                shift = detect_shift(sn, fn)
                # ★ 날짜는 파일명 기준으로만 정리한다.
                #   시트 안 제목(REPORT BY TIME ...) 날짜는 갱신 안 된 채로 남아있는
                #   경우가 많아(담당자가 복붙 후 날짜 텍스트를 안 고침) 신뢰할 수 없음.
                #   파일명이 업로더가 실제로 의도한 날짜이므로 이것만 사용한다.
                ds = fd if fd != "UNKNOWN" else (parse_date(sn) or "UNKNOWN")
                if ds == "UNKNOWN":
                    ds = find_date_in_sheet(ws) or "UNKNOWN"
                if ds == "UNKNOWN": continue

                pkey = (ds, process, shift)
                if pkey in parsed_keys:
                    status.warning(f"중복 스킵: {fn} [{sn}] {ds}/{process}/{shift}")
                    continue
                parsed_keys.add(pkey)

                try: loss_records.extend(parse_sheet(ws, process, ds, shift))
                except Exception as e: st.warning(f"파싱오류[{sn}]: {e}")
        prog.progress((fi+1)/len(uploaded_files))

    status.empty(); prog.empty()
    ldf = pd.DataFrame(loss_records)
    if not ldf.empty and "date" in ldf.columns:
        ldf = ldf[ldf["date"] != "UNKNOWN"].reset_index(drop=True)
    sdf = pd.concat(scrap_list, ignore_index=True) if scrap_list else pd.DataFrame()
    return ldf, sdf

# ─────────────────────────────────────────────
# 유틸
# ─────────────────────────────────────────────
def to_excel(df):
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="xlsxwriter") as w:
        df.to_excel(w, sheet_name="원본데이터", index=False)
        (df.groupby("loss_type_name")["loss_min"].sum().reset_index()
         .sort_values("loss_min", ascending=False)
         .rename(columns={"loss_type_name":"유형","loss_min":"손실(분)"})
         .to_excel(w, sheet_name="유형별", index=False))
        (df.groupby(["process","line"])["loss_min"].sum().reset_index()
         .sort_values("loss_min", ascending=False)
         .rename(columns={"loss_min":"손실(분)"})
         .to_excel(w, sheet_name="라인별", index=False))
    return buf.getvalue()

def to_excel_scrap(df):
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="xlsxwriter") as w:
        (df.groupby("cause_name")["qty"].sum().reset_index()
         .sort_values("qty", ascending=False)
         .rename(columns={"cause_name":"원인","qty":"수량"})
         .to_excel(w, sheet_name="원인별", index=False))
        (df.groupby(["process","line"])["qty"].sum().reset_index()
         .sort_values("qty", ascending=False)
         .rename(columns={"qty":"수량"})
         .to_excel(w, sheet_name="라인별", index=False))
        df.to_excel(w, sheet_name="원본데이터", index=False)
    return buf.getvalue()

def reset_all():
    keys = ["kpi_focus","sel_proc","type_chart","line_chart",
            "proc_pie_chart","scrap_cause_chart","scrap_line_chart"]
    for k in keys:
        if k in st.session_state: del st.session_state[k]

def get_ach_color(v):
    if v >= 100:  return "#16a34a"
    elif v >= 90: return "#475569"
    elif v >= 80: return "#f97316"
    else:         return "#dc2626"

# ─────────────────────────────────────────────
# 로그인
# ─────────────────────────────────────────────
def check_login():
    try: return st.secrets["users"]
    except: return {"hansol": "hevh2024"}

def login_page():
    st.markdown("""
    <div style="text-align:center;margin-top:100px;">
    <h2 style="color:#1e293b;">🏭 한솔테크닉스 HEVH</h2>
    <p style="color:#64748b;">LOSSTIME + SCRAP 분석 대시보드</p>
    </div>""", unsafe_allow_html=True)
    _, col, _ = st.columns([1.5, 1, 1.5])
    with col:
        with st.form("login_form"):
            u = st.text_input("아이디")
            p = st.text_input("비밀번호", type="password")
            if st.form_submit_button("로그인", use_container_width=True, type="primary"):
                users = check_login()
                if u in users and users[u] == p:
                    st.session_state["logged_in"] = True
                    st.session_state["username"]  = u
                    st.rerun()
                else:
                    st.error("아이디 또는 비밀번호 오류")

# ─────────────────────────────────────────────
# 대시보드
# ─────────────────────────────────────────────
def dashboard():
    # ── 헤더
    st.markdown("""
    <div class="main-header">
      <h1>🏭 한솔테크닉스 HEVH — LOSSTIME 분석</h1>
      <p>AI / SMT / PBA(MI) 공정 실적 | 호치민 법인</p>
    </div>""", unsafe_allow_html=True)

    # ★ 자동 DB 로드 (세션에 없을 때만 한 번 실행)
    if "df" not in st.session_state:
        with st.spinner("DB 로드 중..."):
            db  = load_db()
            sdb = load_scrap_db()
        if not db.empty:
            st.session_state["df"] = db
        if not sdb.empty:
            st.session_state["scrap_df"] = sdb

    # ════════════════════════════════════════
    # 사이드바
    # ════════════════════════════════════════
    with st.sidebar:
        st.markdown("### ⚙️ 설정")

        # 파일 업로드
        with st.expander("📂 파일 업로드", expanded=True):
            uploaded = st.file_uploader("xlsx / xls", type=["xlsx","xls"],
                                        accept_multiple_files=True,
                                        label_visibility="collapsed")
            if uploaded:
                if st.button("🚀 분석 / 누적", type="primary", use_container_width=True):
                    with st.spinner("처리 중..."):
                        nl, ns = parse_files(uploaded)
                    if not nl.empty:
                        ex = load_db(); mg = merge_db(ex, nl)
                        with st.spinner("저장..."): save_db(mg)
                        st.session_state["df"] = mg
                        st.success(f"OK {len(nl):,}건")
                    if not ns.empty:
                        es = load_scrap_db(); ms = merge_scrap_db(es, ns)
                        with st.spinner("저장..."): save_scrap_db(ms)
                        st.session_state["scrap_df"] = ms
                        st.success(f"OK SCRAP {len(ns):,}건")
                    if nl.empty and ns.empty: st.warning("데이터 없음 — 스킵")

        if st.button("💾 DB 불러오기", use_container_width=True):
            with st.spinner("로드 중..."):
                db = load_db(); sdb = load_scrap_db()
            if not db.empty:
                st.session_state["df"] = db
                st.success(f"OK {len(db):,}건")
            if not sdb.empty:
                st.session_state["scrap_df"] = sdb
                st.success(f"OK SCRAP {len(sdb):,}건")
            if db.empty and sdb.empty: st.warning("DB 없음")

        st.divider()

        df_all = st.session_state.get("df", pd.DataFrame())
        if df_all.empty:
            st.info("데이터를 불러오세요."); return

        # ── DB 날짜 범위
        dates = sorted([d for d in df_all["date"].dropna().unique() if d != "UNKNOWN"])
        if not dates: st.warning("날짜 없음"); return

        db_min_str = dates[0]
        db_max_str = dates[-1]

        # ★ v5.1: 월별 빠른선택 버튼
        st.markdown("##### 📅 날짜 범위")

        # 존재하는 월 목록 추출
        months_in_db = sorted(set(d[:7] for d in dates))  # ["2026-05", "2026-06", "2026-07"]

        st.markdown("**월 빠른선택**")
        # 월 버튼을 한 행에 나란히
        month_cols = st.columns(len(months_in_db)) if months_in_db else []
        for i, ym in enumerate(months_in_db):
            yr, mo = int(ym[:4]), int(ym[5:7])
            label  = f"{mo}월"
            month_dates = [d for d in dates if d.startswith(ym)]
            if month_cols and month_cols[i].button(label, key=f"mth_{ym}",
                                                    use_container_width=True):
                st.session_state["dr_start"] = month_dates[0]
                st.session_state["dr_end"]   = month_dates[-1]

        # 전체 버튼
        if st.button("전체 기간", key="btn_all", use_container_width=True):
            st.session_state["dr_start"] = db_min_str
            st.session_state["dr_end"]   = db_max_str

        # select_slider (월 버튼 세팅 반영)
        def_s = st.session_state.get("dr_start", db_min_str)
        def_e = st.session_state.get("dr_end",   db_max_str)
        # 범위 보정
        if def_s not in dates: def_s = dates[0]
        if def_e not in dates: def_e = dates[-1]
        if def_s > def_e: def_s = def_e

        if len(dates) >= 2:
            date_range = st.select_slider("", options=dates,
                                          value=(def_s, def_e),
                                          label_visibility="collapsed")
        else:
            date_range = (dates[0], dates[0])

        # 공정
        st.caption("🏭 공정")
        procs = st.multiselect("", ["AI","SMT","MI"],
                               default=["AI","SMT","MI"],
                               label_visibility="collapsed")

        # 주야간
        st.caption("🌙 주야간")
        shifts = st.multiselect("", ["DAY","NIGHT"],
                                default=["DAY","NIGHT"],
                                label_visibility="collapsed")

        with st.expander("⚙️ 추가 필터"):
            lines_all = sorted(df_all["line"].dropna().unique())
            sel_lines = st.multiselect("라인", lines_all)
            types_all = sorted(df_all["loss_type_name"].dropna().unique())
            sel_types = st.multiselect("손실유형", types_all)
            view_mode = st.radio("조회 단위",
                                 ["TOTAL(일계)","타임별(A~K)"],
                                 horizontal=True)

        st.divider()

        # DB 현황
        cd1, cd2 = st.columns(2)
        with cd1:
            dba = load_db()
            if not dba.empty:
                st.info(f"LOSS: {len(dba):,}건\n{dba['date'].nunique()}일 {dba['line'].nunique()}개 라인")
            if st.button("🗑️ LOSSTIME 초기화"):
                github_save_csv(pd.DataFrame(), DB_PATH, "초기화")
                st.session_state.pop("df", None); st.success("완료"); st.rerun()
        with cd2:
            sdba = load_scrap_db()
            if not sdba.empty:
                st.info(f"SCRAP: {len(sdba):,}건\n{sdba['line'].nunique()}개 라인")
            if st.button("🗑️ SCRAP 초기화"):
                github_save_csv(pd.DataFrame(), SCRAP_DB, "초기화")
                st.session_state.pop("scrap_df", None); st.success("완료"); st.rerun()

    # ════════════════════════════════════════
    # 메인 데이터 필터
    # ════════════════════════════════════════
    df       = st.session_state.get("df", pd.DataFrame())
    scrap_df = st.session_state.get("scrap_df", pd.DataFrame())
    if df.empty:
        st.info("👈 파일 업로드 또는 DB 불러오기"); return

    mask = (
        (df["date"] >= date_range[0]) & (df["date"] <= date_range[1]) &
        (df["shift"].isin(shifts or ["DAY","NIGHT"])) &
        (df["process"].isin(procs or ["AI","SMT","MI"]))
    )
    if sel_lines: mask &= df["line"].isin(sel_lines)
    if sel_types: mask &= df["loss_type_name"].isin(sel_types)
    if view_mode == "TOTAL(일계)": mask &= (df["time_slot"] == "TOTAL")
    else:                          mask &= (df["time_slot"] != "TOTAL")

    fdf = df[mask].copy()
    total_df = df[
        (df["date"] >= date_range[0]) & (df["date"] <= date_range[1]) &
        (df["shift"].isin(shifts or ["DAY","NIGHT"])) &
        (df["process"].isin(procs or ["AI","SMT","MI"])) &
        (df["time_slot"] == "TOTAL")
    ].copy()
    if sel_lines: total_df = total_df[total_df["line"].isin(sel_lines)]
    if sel_types: total_df = total_df[total_df["loss_type_name"].isin(sel_types)]

    if fdf.empty:
        st.warning("조건에 맞는 데이터 없음"); return

    # ★ target/actual은 원인(sub_idx)별로 라인 전체 값이 반복 저장되어 있으므로
    #   (date,shift,process,line) 기준으로 한 줄만 남긴 뒤 합산해야 중복 합산을 피할 수 있음
    tgt_key = ["date","shift","process","line"]
    tgt_df  = (total_df.drop_duplicates(subset=tgt_key)
               if not total_df.empty else total_df)

    # KPI
    total_min   = round(total_df["loss_min"].sum(), 1) if not total_df.empty else 0
    total_hr    = round(total_min / 60, 1)
    n_lines     = total_df["line"].nunique() if not total_df.empty else 0
    n_days      = fdf["date"].nunique()
    scrap_total = int(scrap_df["qty"].sum()) if not scrap_df.empty else 0
    t_sum       = tgt_df["target"].sum()
    a_sum       = tgt_df["actual"].sum()
    ach_rate    = round(a_sum / t_sum * 100, 1) if t_sum > 0 else 0

    if "kpi_focus" not in st.session_state:
        st.session_state["kpi_focus"] = None

    # KPI 카드
    k1, k2, k3, k4, k5, k6 = st.columns(6)
    kpi_data = [
        (k1, f"{total_min:,.1f}분", "총 손실", "loss"),
        (k2, f"{total_hr}h",        "손실 시간", None),
        (k3, f"{n_lines}개",        "손실 라인", "line"),
        (k4, f"{n_days}일",         "분석 일수", "date"),
        (k5, f"{scrap_total:,}ea", "스크랩",    "scrap"),
        (k6, f"{ach_rate}%",       "달성률",    "plan"),
    ]
    for col, val, lbl, focus_key in kpi_data:
        cls = "kpi-card"
        if focus_key == "plan":
            color = get_ach_color(ach_rate)
            col.markdown(f"""<div class="{cls}" style="cursor:pointer"
                onclick="">
                <div class="kpi-val" style="color:{color}">{val}</div>
                <div class="kpi-lbl">{lbl}</div></div>""",
                unsafe_allow_html=True)
        else:
            col.markdown(f"""<div class="{cls}">
                <div class="kpi-val">{val}</div>
                <div class="kpi-lbl">{lbl}</div></div>""",
                unsafe_allow_html=True)

    st.divider()

    # ════════════════════════════════════════
    # 탭
    # ════════════════════════════════════════
    tab1,tab2,tab3,tab4,tab5,tab6,tab7,tab8,tab9,tab10 = st.tabs([
        "📊 손실 분석","🏭 라인별","📋 Plan/Actual",
        "📈 트렌드","⏰ 타임별","🍅 스크랩",
        "🔍 상세 조회","🔧 PM",
        "🚨 Shipment Alert","📥 다운로드"
    ])

    # ════════ TAB1 - 손실 분석 ════════
    with tab1:
        if total_df.empty:
            st.warning("TOTAL 데이터 없음")
        else:
            col_trend, col_slot = st.columns(2)
            with col_trend:
                st.markdown("#### 📈 날짜별 손실 트렌드")
                dt = (total_df[total_df["time_slot"] == "TOTAL"]
                      .groupby(["date","process"])["loss_min"].sum().reset_index())
                dt["loss_min"] = dt["loss_min"].round(1)
                # ★ v5.1: string → datetime 변환 (7월 표시 보장)
                dt["date"] = pd.to_datetime(dt["date"])
                fig_t = px.line(dt, x="date", y="loss_min", color="process",
                                color_discrete_map=PROC_COLOR, markers=True, height=320,
                                labels={"loss_min":"손실(분)","date":"날짜","process":"공정"})
                fig_t.update_layout(
                    margin=dict(l=0,r=0,t=10,b=0),
                    yaxis=dict(rangemode="tozero"),
                    xaxis=dict(tickformat="%m/%d"),
                    legend=dict(orientation="h",y=1.05))
                st.plotly_chart(fig_t, use_container_width=True)

            with col_slot:
                st.markdown("#### 🕐 시간대별 손실")
                slot_order = ["A","B","C","D","E","F","G","H","I","J","K"]
                slot_df2 = df[
                    (df["date"] >= date_range[0]) & (df["date"] <= date_range[1]) &
                    (df["shift"].isin(shifts or ["DAY","NIGHT"])) &
                    (df["process"].isin(procs or ["AI","SMT","MI"])) &
                    (df["time_slot"] != "TOTAL")
                ].copy()
                if sel_lines: slot_df2 = slot_df2[slot_df2["line"].isin(sel_lines)]
                if not slot_df2.empty:
                    ss = (slot_df2.groupby(["time_slot","process"])["loss_min"].sum().reset_index())
                    ss["loss_min"] = ss["loss_min"].round(1)
                    ss["time_slot"] = pd.Categorical(ss["time_slot"],
                                                     categories=slot_order, ordered=True)
                    ss = ss.sort_values("time_slot")
                    fig_s = px.bar(ss, x="time_slot", y="loss_min", color="process",
                                   color_discrete_map=PROC_COLOR, barmode="stack", height=320,
                                   labels={"loss_min":"손실(분)","time_slot":"시간대","process":"공정"})
                    fig_s.update_layout(margin=dict(l=0,r=0,t=10,b=0),
                                        yaxis=dict(rangemode="tozero"),
                                        legend=dict(orientation="h",y=1.05))
                    st.plotly_chart(fig_s, use_container_width=True)
                else:
                    st.info("타임별 데이터 없음")

            st.markdown("<hr class='section-divider'>", unsafe_allow_html=True)
            col_l, col_r = st.columns(2)
            with col_l:
                st.markdown("#### 손실유형별 누계")
                # ★ 유형이 20개 넘게 다 나오면 못 읽으니 상위 8개만 표시
                ts = (total_df.groupby("loss_type_name")["loss_min"]
                      .sum().reset_index()
                      .sort_values("loss_min", ascending=False)
                      .head(8))
                ts["loss_min"] = ts["loss_min"].round(1)
                ft = px.bar(ts,
                            x="loss_min",
                            y="loss_type_name",
                            orientation="h",
                            color="loss_type_name",
                            color_discrete_map=TYPE_COLOR,
                            height=500,   # ★ 옆 '라인별 누계 TOP 15' 차트와 높이를 맞춤
                            labels={"loss_min":"손실(분)", "loss_type_name":"유형"})
                ft.update_layout(
                    showlegend=False,
                    margin=dict(l=130, r=20, t=10, b=0),
                    yaxis=dict(categoryorder="total ascending",
                               tickfont=dict(size=12)),
                    xaxis=dict(rangemode="tozero"))
                ct = st.plotly_chart(ft, use_container_width=True,
                                     on_select="rerun", key="type_chart")
                if ct and ct.get("selection",{}).get("points"):
                    sn2 = ct["selection"]["points"][0]
                    sn2 = sn2.get("label") or sn2.get("y","")
                    if sn2:
                        st.markdown(f'<div class="detail-box"><b>📋 {sn2} 상세</b></div>',
                                    unsafe_allow_html=True)
                        dd = fdf[fdf["loss_type_name"] == sn2][
                            ["date","shift","line","time_slot","model",
                             "loss_min","loss_detail"]].sort_values(["date","line"])
                        dd["loss_min"] = dd["loss_min"].round(1)
                        st.dataframe(dd.reset_index(drop=True),
                                     use_container_width=True, height=250)

            with col_r:
                st.markdown("#### 라인별 누계 TOP 15")
                ls = (total_df.groupby(["process","line"])["loss_min"]
                      .sum().reset_index().sort_values("loss_min", ascending=False).head(15))
                ls["loss_min"] = ls["loss_min"].round(1)
                fig2 = px.bar(ls, x="line", y="loss_min", color="process",
                              color_discrete_map=PROC_COLOR, height=500,
                              labels={"loss_min":"손실(분)","line":"라인"})
                fig2.update_layout(margin=dict(l=0,r=0,t=10,b=0),
                                   yaxis=dict(rangemode="tozero"),
                                   legend=dict(orientation="h",y=1.05))
                cl = st.plotly_chart(fig2, use_container_width=True,
                                     on_select="rerun", key="line_chart")
                if cl and cl.get("selection",{}).get("points"):
                    sl = cl["selection"]["points"][0].get("x","")
                    if sl:
                        st.markdown(f'<div class="detail-box"><b>📋 {sl} 라인 상세</b></div>',
                                    unsafe_allow_html=True)
                        ld = fdf[fdf["line"] == sl][
                            ["date","shift","time_slot","model",
                             "loss_min","loss_type_name","loss_detail"]
                        ].sort_values(["date","time_slot"])
                        ld["loss_min"] = ld["loss_min"].round(1)
                        st.dataframe(ld.reset_index(drop=True),
                                     use_container_width=True, height=280)

            st.markdown("<hr class='section-divider'>", unsafe_allow_html=True)
            col_pie, col_type = st.columns(2)
            with col_pie:
                st.markdown("#### 공정별 비중")
                pp = (total_df.groupby("process")["loss_min"].sum().reset_index())
                pp["loss_min"] = pp["loss_min"].round(1)
                fp = px.pie(pp, values="loss_min", names="process",
                            color="process", color_discrete_map=PROC_COLOR, height=350)
                fp.update_traces(textposition="inside", textinfo="percent+label")
                fp.update_layout(margin=dict(l=0,r=0,t=10,b=0), showlegend=False)
                cp = st.plotly_chart(fp, use_container_width=True,
                                     on_select="rerun", key="proc_pie_chart")
                if cp and cp.get("selection",{}).get("points"):
                    sp = cp["selection"]["points"][0].get("label","")
                    if sp: st.session_state["sel_proc"] = sp

                # ★ 파이차트 클릭이 안 먹힐 때를 대비해 버튼으로도 바로 전환 가능하게
                proc_opts = [p for p in ["AI","SMT","MI"] if p in pp["process"].unique()]
                if proc_opts:
                    btn_cols = st.columns(len(proc_opts))
                    for bc, p in zip(btn_cols, proc_opts):
                        is_sel = st.session_state.get("sel_proc", "MI") == p
                        if bc.button(p, key=f"sel_proc_btn_{p}",
                                     type="primary" if is_sel else "secondary",
                                     use_container_width=True):
                            st.session_state["sel_proc"] = p
                            st.rerun()

            with col_type:
                sel_proc = st.session_state.get("sel_proc", "MI")
                # ★ 유형이 많은 공정(SMT 등)은 상위 8개만 표시
                proc_type = (total_df[total_df["process"] == sel_proc]
                             .groupby("loss_type_name")["loss_min"].sum()
                             .reset_index()
                             .sort_values("loss_min", ascending=False)
                             .head(8)
                             .sort_values("loss_min", ascending=True))
                proc_type["loss_min"] = proc_type["loss_min"].round(1)
                st.markdown(f"**{sel_proc} 손실유형 TOP 8**")
                ft = px.bar(proc_type,
                            x="loss_min",
                            y="loss_type_name",
                            orientation="h",
                            color="loss_type_name",
                            color_discrete_map=TYPE_COLOR,
                            height=350,   # ★ 옆 파이차트(350)와 높이를 맞춤
                            labels={"loss_min":"손실(분)", "loss_type_name":"유형"})
                ft.update_layout(
                    showlegend=False,
                    margin=dict(l=130, r=20, t=10, b=0),
                    xaxis=dict(rangemode="tozero"),
                    yaxis=dict(categoryorder="total ascending",
                               tickfont=dict(size=12))
                )
                st.plotly_chart(ft, use_container_width=True)

            st.markdown("<hr class='section-divider'>", unsafe_allow_html=True)
            col_trend2, col_slot2 = st.columns(2)
            with col_trend2:
                st.markdown("#### 날짜별 트렌드 (라인별)")
                dt2 = (total_df.groupby(["date","shift"])["loss_min"].sum().reset_index())
                dt2["loss_min"] = dt2["loss_min"].round(1)
                dt2["date"] = pd.to_datetime(dt2["date"])  # ★ v5.1
                fig_d2 = px.bar(dt2, x="date", y="loss_min", color="shift",
                                color_discrete_map={"DAY":"#f59e0b","NIGHT":"#6366f1"},
                                height=300, barmode="stack",
                                labels={"loss_min":"손실(분)","date":"날짜","shift":"구분"})
                fig_d2.update_layout(margin=dict(l=0,r=0,t=10,b=0),
                                     yaxis=dict(rangemode="tozero"),
                                     xaxis=dict(tickformat="%m/%d"),
                                     legend=dict(orientation="h",y=1.05))
                st.plotly_chart(fig_d2, use_container_width=True)

            with col_slot2:
                st.markdown("#### DAY vs NIGHT")
                sc = (total_df.groupby(["shift","loss_type_name"])["loss_min"].sum().reset_index())
                sc["loss_min"] = sc["loss_min"].round(1)
                fig5 = px.bar(sc, x="loss_type_name", y="loss_min", color="shift",
                              color_discrete_map={"DAY":"#f59e0b","NIGHT":"#6366f1"},
                              barmode="group", height=300,
                              labels={"loss_min":"손실(분)","loss_type_name":"손실유형","shift":"구분"})
                fig5.update_layout(margin=dict(l=0,r=0,t=10,b=0),
                                   xaxis_tickangle=-30, yaxis=dict(rangemode="tozero"))
                st.plotly_chart(fig5, use_container_width=True)

    # ════════ TAB2 - 라인별 ════════
    with tab2:
        if total_df.empty:
            st.warning("TOTAL 데이터 없음")
        else:
            st.markdown("#### 🏭 라인별 상세 분석")
            proc_sel = st.radio("공정 선택", ["전체","AI","SMT","MI"],
                                horizontal=True, key="line_tab_proc")
            line_df = total_df.copy() if proc_sel == "전체" else \
                      total_df[total_df["process"] == proc_sel].copy()
            line_df = line_df[line_df["time_slot"] == "TOTAL"].copy()
            if line_df.empty:
                st.warning("데이터 없음")
            else:
                tot_proc = round(line_df["loss_min"].sum(), 1)
                c1, c2, c3 = st.columns(3)
                c1.metric("📦 공정 전체 누계 손실", f"{tot_proc:,.1f}분")
                c2.metric("⏱ 손실 시간", f"{round(tot_proc/60,1)}h")
                c3.metric("📅 발생 일수", f"{line_df['date'].nunique()}일")

                st.markdown("<br>", unsafe_allow_html=True)

                lt_proc = (line_df.groupby("loss_type_name")["loss_min"]
                           .sum().reset_index().sort_values("loss_min", ascending=False))
                lt_proc["loss_min"] = lt_proc["loss_min"].round(1)
                total_lt = lt_proc["loss_min"].sum()
                lt_proc["cum_pct"] = lt_proc["loss_min"].cumsum() / total_lt * 100 if total_lt else 0

                # ★ 유형이 20개 넘게 뒤섞여 있으면 막대만 봐서는 인사이트가 안 나와서,
                #   상위 N개 + "기타(그외 묶음)"로 압축하고 파레토 누적선을 같이 보여준다.
                TOP_N = 8
                top_names = lt_proc.head(TOP_N)["loss_type_name"].tolist()
                disp = lt_proc.head(TOP_N)[["loss_type_name","loss_min"]].copy()
                rest = lt_proc.iloc[TOP_N:]
                if not rest.empty:
                    disp = pd.concat([disp, pd.DataFrame([{
                        "loss_type_name": f"기타(그외 {len(rest)}종)",
                        "loss_min": round(rest["loss_min"].sum(), 1)
                    }])], ignore_index=True)
                disp["cum_pct"] = round(disp["loss_min"].cumsum() / total_lt * 100, 1) if total_lt else 0

                # 80% 지점까지 몇 개 유형이 필요한지 → 핵심 인사이트 문구
                n80 = int((lt_proc["cum_pct"] <= 80).sum()) + 1
                n80 = min(n80, len(lt_proc))
                top_for_insight = lt_proc.head(n80)
                pct80 = top_for_insight["loss_min"].sum() / total_lt * 100 if total_lt else 0
                names_str = ", ".join(top_for_insight["loss_type_name"].tolist())
                st.markdown(
                    f'<div class="detail-box">📌 <b>{n80}개 유형</b>이 전체 손실의 '
                    f'<b>{pct80:.0f}%</b>를 차지합니다: {names_str}</div>',
                    unsafe_allow_html=True
                )

                st.markdown("##### 손실 유형별 집계 — 상위 유형 + 누적비중(파레토)")
                fig_lt = go.Figure()
                fig_lt.add_bar(x=disp["loss_type_name"], y=disp["loss_min"],
                               marker_color=[TYPE_COLOR.get(n, "#94a3b8") for n in disp["loss_type_name"]],
                               name="손실(분)", text=disp["loss_min"],
                               texttemplate="%{text:,.0f}", textposition="outside")
                fig_lt.add_scatter(x=disp["loss_type_name"], y=disp["cum_pct"],
                                    name="누적비중(%)", mode="lines+markers",
                                    line=dict(color="#dc2626", width=2), yaxis="y2")
                fig_lt.update_layout(
                    height=440, margin=dict(l=20, r=40, t=20, b=100),
                    xaxis=dict(tickangle=-30),
                    yaxis=dict(title="손실(분)", rangemode="tozero"),
                    yaxis2=dict(title="누적비중(%)", overlaying="y", side="right",
                                range=[0, 105]),
                    legend=dict(orientation="h", y=1.1), showlegend=True,
                )
                cl_lt = st.plotly_chart(fig_lt, use_container_width=True,
                                        on_select="rerun", key="lt_tab_chart")
                sel_type = ""
                if cl_lt and cl_lt.get("selection",{}).get("points"):
                    sel_type = cl_lt["selection"]["points"][0].get("x","")

                # ★ 막대 클릭 시 해당 손실유형(또는 "기타" 묶음)의 상세 내역을 바로 보여줌
                if sel_type:
                    st.markdown(f'<div class="detail-box"><b>📋 {sel_type} 상세</b></div>',
                                unsafe_allow_html=True)
                    if sel_type.startswith("기타(그외"):
                        dd_lt = line_df[~line_df["loss_type_name"].isin(top_names)]
                    else:
                        dd_lt = line_df[line_df["loss_type_name"] == sel_type]
                    dd_lt = dd_lt[["date","shift","process","line","loss_min",
                                    "loss_type_name","loss_detail","action"]].sort_values(
                                    ["date","line"])
                    dd_lt["loss_min"] = dd_lt["loss_min"].round(1)
                    st.dataframe(dd_lt.reset_index(drop=True),
                                 use_container_width=True, height=280)

                st.markdown("<hr class='section-divider'>", unsafe_allow_html=True)
                all_lines = sorted(line_df["line"].dropna().unique())
                col_sel, _ = st.columns([1,3])
                with col_sel:
                    chosen = st.selectbox("🔍 라인 선택", all_lines, key="line_tab_sel")

                if chosen:
                    ld = total_df[total_df["line"] == chosen].copy()
                    total_line = round(ld["loss_min"].sum(), 1)
                    m1,m2,m3,m4 = st.columns(4)
                    for col_m, val_m, lbl_m in [
                        (m1, f"{total_line:,.1f}분", "누계 손실"),
                        (m2, f"{round(total_line/60,1)}h", "손실 시간"),
                        (m3, f"{ld['date'].nunique()}일", "발생 일수"),
                        (m4, f"{ld['loss_type_name'].nunique()}종", "손실 유형수"),
                    ]:
                        col_m.markdown(f"""<div class="kpi-card">
                            <div class="kpi-val">{val_m}</div>
                            <div class="kpi-lbl">{lbl_m}</div></div>""",
                            unsafe_allow_html=True)

                    st.markdown("<br>", unsafe_allow_html=True)
                    col_a, col_b = st.columns(2)
                    with col_a:
                        st.markdown(f"##### {chosen} - 손실유형 비중")
                        lt2 = (ld.groupby("loss_type_name")["loss_min"].sum().reset_index()
                               .sort_values("loss_min", ascending=False))
                        lt2["loss_min"] = lt2["loss_min"].round(1)
                        fp2 = px.pie(lt2, values="loss_min", names="loss_type_name",
                                     color="loss_type_name", color_discrete_map=TYPE_COLOR,
                                     height=300)
                        fp2.update_traces(textposition="inside", textinfo="percent+label")
                        fp2.update_layout(margin=dict(l=0,r=0,t=10,b=0), showlegend=False)
                        st.plotly_chart(fp2, use_container_width=True)
                    with col_b:
                        st.markdown(f"##### {chosen} - 날짜별 트렌드")
                        dt_l = (ld.groupby(["date","shift"])["loss_min"].sum().reset_index())
                        dt_l["loss_min"] = dt_l["loss_min"].round(1)
                        dt_l["date"] = pd.to_datetime(dt_l["date"])  # ★ v5.1
                        fig_dt = px.bar(dt_l, x="date", y="loss_min", color="shift",
                                        color_discrete_map={"DAY":"#f59e0b","NIGHT":"#6366f1"},
                                        height=300, barmode="stack",
                                        labels={"loss_min":"손실(분)","date":"날짜","shift":"구분"})
                        fig_dt.update_layout(margin=dict(l=0,r=0,t=10,b=0),
                                             yaxis=dict(rangemode="tozero"),
                                             xaxis=dict(tickformat="%m/%d"),
                                             legend=dict(orientation="h",y=1.05))
                        st.plotly_chart(fig_dt, use_container_width=True)

    # ════════ TAB3 - Plan/Actual ════════
    with tab3:
        if total_df.empty:
            st.warning("TOTAL 데이터 없음")
        else:
            st.markdown("#### 📋 Plan / Actual")
            proc_pa = st.radio("공정",["전체","AI","SMT","MI"],horizontal=True,key="pa_proc")
            pa_df = total_df.copy() if proc_pa == "전체" else \
                    total_df[total_df["process"] == proc_pa].copy()
            # ★ 위와 동일한 이유로 라인당 한 줄만 남긴 뒤 target/actual 합산
            pa_tgt = (pa_df.drop_duplicates(subset=["date","shift","process","line"])
                      if not pa_df.empty else pa_df)

            if proc_pa == "MI":
                t_sum2  = pa_tgt["target_mi"].sum()
                a_sum2  = pa_tgt["actual_mi"].sum()
            else:
                t_sum2 = pa_tgt["target"].sum()
                a_sum2 = pa_tgt["actual"].sum()
            gap2  = a_sum2 - t_sum2
            ach2  = round(a_sum2 / t_sum2 * 100, 1) if t_sum2 > 0 else 0

            c1,c2,c3,c4 = st.columns(4)
            for col_p, val_p, lbl_p, gap_val in [
                (c1, f"{int(t_sum2):,}", "TARGET", None),
                (c2, f"{int(a_sum2):,}", "ACTUAL", None),
                (c3, f"{int(gap2):+,}",  "GAP",   gap2),
                (c4, f"{ach2}%",         "달성률", ach2-100),
            ]:
                if gap_val is not None:
                    cls  = "kpi-card-green" if gap_val >= 0 else "kpi-card-red"
                    vcls = "kpi-val-green"  if gap_val >= 0 else "kpi-val-red"
                else:
                    cls = "kpi-card"; vcls = "kpi-val"
                col_p.markdown(f"""<div class="{cls}">
                    <div class="{vcls}">{val_p}</div>
                    <div class="kpi-lbl">{lbl_p}</div></div>""",
                    unsafe_allow_html=True)

            st.markdown("##### 날짜별 Target vs Actual")
            if proc_pa == "MI":
                dt_pa = (pa_tgt.groupby("date")[["target_mi","actual_mi",
                                                 "target_ate","actual_ate"]].sum().reset_index())
                fig_pa = go.Figure()
                fig_pa.add_scatter(x=pd.to_datetime(dt_pa["date"]),y=dt_pa["target_mi"],
                                   name="MI Target",line=dict(dash="dash",color="#94a3b8"))
                fig_pa.add_scatter(x=pd.to_datetime(dt_pa["date"]),y=dt_pa["actual_mi"],
                                   name="MI Actual",line=dict(color="#10b981",width=2),
                                   mode="lines+markers")
                fig_pa.add_scatter(x=pd.to_datetime(dt_pa["date"]),y=dt_pa["target_ate"],
                                   name="ATE Target",line=dict(dash="dot",color="#cbd5e1"))
                fig_pa.add_scatter(x=pd.to_datetime(dt_pa["date"]),y=dt_pa["actual_ate"],
                                   name="ATE Actual",line=dict(color="#3b82f6",width=2),
                                   mode="lines+markers")
            else:
                dt_pa = (pa_tgt.groupby(["date","process"])[["target","actual"]].sum().reset_index())
                fig_pa = go.Figure()
                for proc in dt_pa["process"].unique():
                    sub = dt_pa[dt_pa["process"] == proc]
                    c   = PROC_COLOR.get(proc,"#94a3b8")
                    fig_pa.add_scatter(x=pd.to_datetime(sub["date"]),y=sub["target"],
                                       name=f"{proc} Target",
                                       line=dict(dash="dash",color=c),opacity=0.5)
                    fig_pa.add_scatter(x=pd.to_datetime(sub["date"]),y=sub["actual"],
                                       name=f"{proc} Actual",
                                       line=dict(color=c,width=2),mode="lines+markers")
            fig_pa.update_layout(margin=dict(l=0,r=0,t=10,b=0),
                                 yaxis=dict(rangemode="tozero"),height=360,
                                 xaxis=dict(tickformat="%m/%d"),
                                 legend=dict(orientation="h",y=1.05),
                                 plot_bgcolor="#f8fafc")
            st.plotly_chart(fig_pa, use_container_width=True)

    # ════════ TAB4 - 트렌드 ════════
    with tab4:
        if total_df.empty:
            st.warning("TOTAL 데이터 없음")
        else:
            st.markdown("#### 날짜별 손실 트렌드")
            dt = (total_df.groupby(["date","process"])["loss_min"].sum().reset_index())
            dt["loss_min"] = dt["loss_min"].round(1)
            # ★ v5.1: datetime 변환 → 7월 데이터 정상 표시
            dt["date"] = pd.to_datetime(dt["date"])
            fig4 = px.line(dt, x="date", y="loss_min", color="process",
                           color_discrete_map=PROC_COLOR, markers=True, height=380,
                           labels={"loss_min":"손실(분)","date":"날짜","process":"공정"})
            fig4.update_layout(
                margin=dict(l=0,r=0,t=10,b=0),
                yaxis=dict(rangemode="tozero"),
                xaxis=dict(tickformat="%m/%d",
                           rangeslider=dict(visible=True)),  # ★ 슬라이더 추가
                legend=dict(orientation="h",y=1.05))
            st.plotly_chart(fig4, use_container_width=True)

            st.markdown("#### 손실유형 트렌드")
            dt2 = (total_df.groupby(["date","loss_type_name"])["loss_min"].sum().reset_index())
            dt2["loss_min"] = dt2["loss_min"].round(1)
            dt2["date"] = pd.to_datetime(dt2["date"])  # ★ v5.1
            top_types = (dt2.groupby("loss_type_name")["loss_min"]
                         .sum().nlargest(8).index.tolist())
            dt2_top = dt2[dt2["loss_type_name"].isin(top_types)]
            fig4b = px.line(dt2_top, x="date", y="loss_min", color="loss_type_name",
                            color_discrete_map=TYPE_COLOR, markers=True, height=360,
                            labels={"loss_min":"손실(분)","date":"날짜","loss_type_name":"유형"})
            fig4b.update_layout(margin=dict(l=0,r=0,t=10,b=0),
                                yaxis=dict(rangemode="tozero"),
                                xaxis=dict(tickformat="%m/%d"),
                                legend=dict(orientation="h",y=1.05))
            st.plotly_chart(fig4b, use_container_width=True)

    # ════════ TAB5 - 타임별 ════════
    with tab5:
        st.markdown("#### 라인 x 시간대 히트맵")
        slot_order = ["A","B","C","D","E","F","G","H","I","J","K"]
        slot_df = df[
            (df["date"] >= date_range[0]) & (df["date"] <= date_range[1]) &
            (df["shift"].isin(shifts or ["DAY","NIGHT"])) &
            (df["process"].isin(procs or ["AI","SMT","MI"])) &
            (df["time_slot"] != "TOTAL")
        ].copy()
        if sel_lines: slot_df = slot_df[slot_df["line"].isin(sel_lines)]
        if slot_df.empty:
            st.warning("타임별 데이터 없음")
        else:
            pivot = (slot_df.groupby(["line","time_slot"])["loss_min"].sum().reset_index())
            pt = pivot.pivot(index="line", columns="time_slot", values="loss_min").fillna(0).round(1)
            pt = pt[[c for c in slot_order if c in pt.columns]]
            # ★ 손실 큰 라인 순으로 정렬하고, 화면을 다 차지하지 않도록 기본은 상위 20개만
            #   표시 (라인이 30개 넘으면 한 화면에 다 넣어도 어차피 안 읽힘).
            pt = pt.loc[pt.sum(axis=1).sort_values(ascending=False).index]
            show_all_lines = False
            if len(pt) > 20:
                show_all_lines = st.checkbox(
                    f"전체 {len(pt)}개 라인 다 보기 (기본: 손실 큰 상위 20개만)",
                    key="heatmap_show_all")
            pt_disp = pt if show_all_lines else pt.head(20)
            fh = px.imshow(pt_disp, color_continuous_scale="Reds", aspect="auto",
                           height=min(650, max(320, len(pt_disp)*24)),
                           labels={"x":"시간대","y":"라인","color":"손실(분)"})
            fh.update_layout(margin=dict(l=0,r=0,t=10,b=0),
                              yaxis=dict(categoryorder="array",
                                         categoryarray=pt_disp.index[::-1]))
            st.plotly_chart(fh, use_container_width=True)

            st.markdown("#### 시간대별 합계")
            ss = (slot_df.groupby(["time_slot","process"])["loss_min"].sum().reset_index())
            ss["loss_min"] = ss["loss_min"].round(1)
            ss["time_slot"] = pd.Categorical(ss["time_slot"], categories=slot_order, ordered=True)
            ss = ss.sort_values("time_slot")
            fs = px.bar(ss, x="time_slot", y="loss_min", color="process",
                        color_discrete_map=PROC_COLOR, barmode="stack", height=320,
                        labels={"loss_min":"손실(분)","time_slot":"시간대","process":"공정"})
            fs.update_layout(margin=dict(l=0,r=0,t=10,b=0), yaxis=dict(rangemode="tozero"))
            st.plotly_chart(fs, use_container_width=True)

            # ★ 시간대별로 어떤 손실유형이 큰지 바로 보이도록 유형별 누적 막대 추가
            st.markdown("#### 시간대별 손실유형 집계")
            st_type = (slot_df.groupby(["time_slot","loss_type_name"])["loss_min"]
                       .sum().reset_index())
            st_type["loss_min"] = st_type["loss_min"].round(1)
            st_type["time_slot"] = pd.Categorical(st_type["time_slot"],
                                                   categories=slot_order, ordered=True)
            st_type = st_type.sort_values("time_slot")
            fst = px.bar(st_type, x="time_slot", y="loss_min", color="loss_type_name",
                         color_discrete_map=TYPE_COLOR, barmode="stack", height=380,
                         labels={"loss_min":"손실(분)","time_slot":"시간대","loss_type_name":"유형"})
            fst.update_layout(margin=dict(l=0,r=0,t=10,b=0), yaxis=dict(rangemode="tozero"),
                              legend=dict(orientation="h", y=-0.25))
            st.plotly_chart(fst, use_container_width=True)

            # 손실이 가장 큰 시간대 상위 3개는 원인 TOP5를 바로 요약해서 보여줌
            top_slots = (slot_df.groupby("time_slot")["loss_min"].sum()
                         .sort_values(ascending=False).head(3))
            st.markdown("##### 🔎 손실 최다 시간대 원인 TOP5")
            cols_top = st.columns(len(top_slots))
            for col_ts, (slot_name, slot_total) in zip(cols_top, top_slots.items()):
                with col_ts:
                    st.markdown(f"**{slot_name}시간대 — 총 {slot_total:,.0f}분**")
                    top5 = (slot_df[slot_df["time_slot"] == slot_name]
                            .groupby("loss_type_name")["loss_min"].sum()
                            .sort_values(ascending=False).head(5))
                    for tname, tmin in top5.items():
                        pct = tmin / slot_total * 100 if slot_total else 0
                        st.markdown(f"- {tname}: {tmin:,.0f}분 ({pct:.0f}%)")

            st.markdown("#### 타임별 상세")
            sd = (slot_df.groupby(["date","shift","line","time_slot",
                                    "loss_type_name","loss_detail"])["loss_min"]
                  .sum().reset_index().sort_values(["date","line","time_slot"]))
            sd["loss_min"] = sd["loss_min"].round(1)
            st.dataframe(sd, use_container_width=True, height=360)

    # ════════ TAB6 - 스크랩 ════════
    with tab6:
        st.markdown("#### 📛 스크랩 분석")
        if scrap_df.empty:
            st.info("스크랩 데이터 없음")
        else:
            sd_dates = sorted(scrap_df["work_date"].dropna().unique())
            if len(sd_dates) >= 2:
                sr = st.select_slider("날짜", options=sd_dates,
                                      value=(sd_dates[0], sd_dates[-1]), key="scrap_date")
                sdf2 = scrap_df[(scrap_df["work_date"] >= sr[0]) &
                                 (scrap_df["work_date"] <= sr[1])].copy()
            else:
                sdf2 = scrap_df.copy()

            show_auto = st.checkbox("자동판정(ATE) 포함", value=True)
            if not show_auto: sdf2 = sdf2[sdf2["is_auto"] == "N"]

            if sdf2.empty:
                st.warning("조건에 맞는 스크랩 없음")
            else:
                k1,k2,k3,k4 = st.columns(4)
                for col_s, val_s, lbl_s in [
                    (k1, f"{int(sdf2['qty'].sum()):,}ea", "총 스크랩"),
                    (k2, f"{len(sdf2):,}건", "이력"),
                    (k3, f"{sdf2['model'].nunique()}종", "모델"),
                    (k4, f"{sdf2['line'].nunique()}개", "발생 라인"),
                ]:
                    col_s.markdown(f"""<div class="kpi-card">
                        <div class="kpi-val">{val_s}</div>
                        <div class="kpi-lbl">{lbl_s}</div></div>""",
                        unsafe_allow_html=True)

                st.markdown("<hr class='section-divider'>", unsafe_allow_html=True)
                ca, cb = st.columns(2)
                with ca:
                    st.markdown("#### 원인별 파레토")
                    cg = (sdf2.groupby("cause_name")["qty"].sum().reset_index()
                          .sort_values("qty", ascending=False))
                    cg["누계%"] = (cg["qty"].cumsum() / cg["qty"].sum() * 100).round(1)
                    fp = go.Figure()
                    fp.add_bar(x=cg["cause_name"], y=cg["qty"], name="수량",
                               marker_color="#ef4444")
                    fp.add_scatter(x=cg["cause_name"], y=cg["누계%"],
                                   mode="lines+markers", name="누계%", yaxis="y2",
                                   line=dict(color="#1e3a5f",width=2))
                    fp.update_layout(
                        yaxis2=dict(overlaying="y",side="right",range=[0,110],ticksuffix="%"),
                        yaxis=dict(rangemode="tozero"), height=360,
                        margin=dict(l=0,r=0,t=10,b=0), xaxis_tickangle=-30)
                    cc2 = st.plotly_chart(fp, use_container_width=True,
                                          on_select="rerun", key="scrap_cause_chart")
                    if cc2 and cc2.get("selection",{}).get("points"):
                        sc2 = cc2["selection"]["points"][0].get("x","")
                        if sc2:
                            st.markdown(f"**📋 [{sc2}] 상세**")
                            cd = sdf2[sdf2["cause_name"] == sc2][
                                ["work_date","process","line","model","qty","comment"]
                            ].sort_values("work_date")
                            st.dataframe(cd.reset_index(drop=True),
                                         use_container_width=True, height=200)
                with cb:
                    st.markdown("#### 라인별 TOP 15")
                    lg = (sdf2.groupby(["process","line"])["qty"].sum().reset_index()
                          .sort_values("qty", ascending=False).head(15))
                    fl = px.bar(lg, x="line", y="qty", color="process",
                                color_discrete_map=PROC_COLOR, height=320,
                                labels={"qty":"수량","line":"라인","process":"공정"})
                    fl.update_layout(margin=dict(l=0,r=0,t=10,b=0),
                                     yaxis=dict(rangemode="tozero"),
                                     legend=dict(orientation="h",y=1.05))
                    csl = st.plotly_chart(fl, use_container_width=True,
                                          on_select="rerun", key="scrap_line_chart")
                    if csl and csl.get("selection",{}).get("points"):
                        ssl = csl["selection"]["points"][0].get("x","")
                        if ssl:
                            st.markdown(f"**📋 [{ssl}] 라인 상세**")
                            ld2 = sdf2[sdf2["line"] == ssl][
                                ["work_date","model","qty","cause_name","comment"]
                            ].sort_values("work_date")
                            st.dataframe(ld2.reset_index(drop=True),
                                         use_container_width=True, height=220)

                st.markdown("#### 날짜별 트렌드")
                dg = sdf2.groupby(["work_date","process"])["qty"].sum().reset_index()
                fdt = px.line(dg, x="work_date", y="qty", color="process",
                              color_discrete_map=PROC_COLOR, markers=True, height=300,
                              labels={"qty":"수량","work_date":"날짜","process":"공정"})
                fdt.update_layout(margin=dict(l=0,r=0,t=10,b=0), yaxis=dict(rangemode="tozero"))
                st.plotly_chart(fdt, use_container_width=True)

    # ════════ TAB7 - 상세 조회 ════════
    with tab7:
        st.markdown("#### 🔍 상세 데이터 조회")
        t7c1, t7c2, t7c3, t7c4 = st.columns(4)
        with t7c1:
            t7_proc = st.selectbox("공정", ["전체","AI","SMT","MI"], key="t7_proc")
        with t7c2:
            t7_shift = st.selectbox("주야간", ["전체","DAY","NIGHT"], key="t7_shift")
        with t7c3:
            t7_slot = st.radio("조회 단위", ["TOTAL(일계)","타임별(A~K)"],
                               horizontal=True, key="t7_slot")
        with t7c4:
            all_dates_t7 = sorted([d for d in df["date"].dropna().unique() if d != "UNKNOWN"])
            if len(all_dates_t7) >= 2:
                t7_date = st.select_slider("날짜", options=all_dates_t7,
                                           value=(all_dates_t7[0], all_dates_t7[-1]),
                                           key="t7_date")
            else:
                t7_date = (all_dates_t7[0], all_dates_t7[0]) if all_dates_t7 else (None, None)

        srch2 = st.text_input("🔍 키워드 (라인/원인/모델)", key="t7_srch")

        sdf3 = df.copy()
        if t7_proc != "전체":  sdf3 = sdf3[sdf3["process"] == t7_proc]
        if t7_shift != "전체": sdf3 = sdf3[sdf3["shift"] == t7_shift]
        if t7_slot == "TOTAL(일계)": sdf3 = sdf3[sdf3["time_slot"] == "TOTAL"]
        else:                         sdf3 = sdf3[sdf3["time_slot"] != "TOTAL"]
        if t7_date and t7_date[0]:
            sdf3 = sdf3[(sdf3["date"] >= t7_date[0]) & (sdf3["date"] <= t7_date[1])]
        if srch2:
            m2 = (sdf3["line"].astype(str).str.contains(srch2,case=False,na=False) |
                  sdf3["loss_detail"].astype(str).str.contains(srch2,case=False,na=False) |
                  sdf3["model"].astype(str).str.contains(srch2,case=False,na=False))
            sdf3 = sdf3[m2]

        st.caption(f"총 {len(sdf3):,}건")
        dc2 = ["date","shift","process","line","time_slot",
               "model","loss_min","loss_type_name","complexity","loss_detail"]
        dc2 = [c for c in dc2 if c in sdf3.columns]
        if not sdf3.empty:
            sdf3["loss_min"] = sdf3["loss_min"].round(1)
            st.dataframe(sdf3[dc2].reset_index(drop=True),
                         use_container_width=True, height=500)
        else:
            st.info("검색 결과 없음")

    # ════════ TAB8 - PM ════════
    with tab8:
        st.markdown("#### 🔧 설비보전 PM 우선순위")
        if total_df.empty:
            st.warning("데이터 없음")
        else:
            pm_types = ["Printer불량","Axial불량","RH3삽입불량","Mouter불량",
                        "XGZ불량","설비고장(기타)","Silicon/Coating불량",
                        "AOI/S-AOI불량","Wave Solder불량"]
            pm = (total_df[total_df["loss_type_name"].isin(pm_types)]
                  .groupby(["process","line","loss_type_name"])["loss_min"]
                  .agg(["sum","count"]).reset_index()
                  .sort_values("sum", ascending=False)
                  .rename(columns={"sum":"누계손실(분)","count":"발생횟수"}))
            pm["누계손실(분)"] = pm["누계손실(분)"].round(1)

            def pmg(row):
                if row["발생횟수"] >= 5 or row["누계손실(분)"] >= 500: return "P1 즉시"
                elif row["발생횟수"] >= 3 or row["누계손실(분)"] >= 200: return "P1"
                elif row["발생횟수"] >= 2 or row["누계손실(분)"] >= 100: return "P2"
                else: return "P3"

            if pm.empty:
                st.info("설비 불량 없음")
            else:
                pm["PM등급"] = pm.apply(pmg, axis=1)
                for _, row in pm.iterrows():
                    g   = str(row["PM등급"])
                    cls = ("pm-p1" if "P1" in g else "pm-p2" if "P2" in g else "pm-p3")
                    st.markdown(f"""<div class="{cls}">
                        <b>{g}</b> | {row['process']} - {row['line']} - {row['loss_type_name']}
                        | <b>{row['누계손실(분)']:,.1f}분</b> / {int(row['발생횟수'])}회
                        </div>""", unsafe_allow_html=True)

    # ════════ TAB9 - Shipment Alert ════════
    with tab9:
        render_shipment_alert_tab()

    # ════════ TAB10 - 다운로드 ════════
    with tab10:
        st.markdown("#### 다운로드")
        ca2, cb2, cc2 = st.columns(3)
        with ca2:
            st.download_button("📥 LOSSTIME 엑셀", data=to_excel(fdf),
                               file_name="HEVH_LOSSTIME.xlsx",
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                               use_container_width=True)
        with cb2:
            if not scrap_df.empty:
                st.download_button("📥 SCRAP 엑셀", data=to_excel_scrap(scrap_df),
                                   file_name="HEVH_SCRAP.xlsx",
                                   mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                   use_container_width=True)
            else:
                st.info("스크랩 없음")
        with cc2:
            st.download_button("📥 CSV",
                               data=fdf.to_csv(index=False,encoding="utf-8-sig").encode("utf-8-sig"),
                               file_name="HEVH_LOSSTIME.csv", mime="text/csv",
                               use_container_width=True)

# ─────────────────────────────────────────────
# 진입점
# ─────────────────────────────────────────────
def main():
    if "logged_in" not in st.session_state:
        st.session_state["logged_in"] = False
    if not st.session_state["logged_in"]: login_page()
    else: dashboard()

if __name__ == "__main__":
    main()
